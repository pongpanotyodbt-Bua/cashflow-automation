"""
Build all Excel templates for Cashflow Automation system.
Creates 11 template files in templates/ directory.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from datetime import datetime, timedelta
import os

# === Styling constants ===
FONT_NAME = "Arial"
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill("solid", start_color="2A6FF0")
SUB_HEADER_FONT = Font(name=FONT_NAME, bold=True, size=10)
SUB_HEADER_FILL = PatternFill("solid", start_color="EAF1FE")
SAMPLE_FONT = Font(name=FONT_NAME, color="0000FF", italic=True, size=10)  # blue = input
NORMAL_FONT = Font(name=FONT_NAME, size=10)
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=14, color="2A6FF0")
NOTE_FONT = Font(name=FONT_NAME, italic=True, color="666666", size=9)

THIN = Side(border_style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")

ENTITIES_NOTE = "ACG / HMW / CLIK / CONSO"
SEGMENT_NOTE_IN = "HMW: HONDA/TOK/INS/FIN/RENT  |  ACG: MGT_HMW/MGT_CLIK  |  CLIK: SVC"
SEGMENT_NOTE_OUT = "F1=Loan&Int  I1=Fixed asset  O1=Staff  O2=Inventory  O3=Subcon  O4=Utility  O5=Tax  O6=Other"

OUT_DIR = os.path.dirname(os.path.abspath(__file__))


def setup_sheet(ws, title, columns, sample_rows, notes=None, col_widths=None):
    """Standard template sheet setup with title, headers, sample data, notes."""
    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
    ws.cell(row=1, column=1, value=title).font = TITLE_FONT
    ws.cell(row=1, column=1).alignment = CENTER
    ws.row_dimensions[1].height = 28

    # Notes / instructions (rows 2-4)
    if notes:
        for i, note in enumerate(notes):
            ws.merge_cells(start_row=2+i, start_column=1, end_row=2+i, end_column=len(columns))
            c = ws.cell(row=2+i, column=1, value=note)
            c.font = NOTE_FONT
            c.alignment = LEFT
        header_row = 2 + len(notes) + 1
    else:
        header_row = 3

    # Header row
    for j, col in enumerate(columns, start=1):
        c = ws.cell(row=header_row, column=j, value=col)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = BORDER
    ws.row_dimensions[header_row].height = 30

    # Sample data rows (blue, italic - signals "edit me")
    for r_offset, row in enumerate(sample_rows, start=1):
        for j, val in enumerate(row, start=1):
            c = ws.cell(row=header_row + r_offset, column=j, value=val)
            c.font = SAMPLE_FONT
            c.alignment = LEFT if isinstance(val, str) else RIGHT
            c.border = BORDER
            if isinstance(val, (int, float)) and not isinstance(val, bool):
                c.number_format = '#,##0.00'
            elif isinstance(val, datetime):
                c.number_format = 'YYYY-MM-DD'
                c.alignment = CENTER

    # Column widths
    if col_widths:
        for i, w in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
    else:
        for i in range(1, len(columns) + 1):
            ws.column_dimensions[get_column_letter(i)].width = 16

    # Freeze first column + header
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)


def add_instructions_sheet(wb, title, purpose, usage_steps, columns_doc, notes=None):
    """Create an Instructions sheet (always sheet index 0)."""
    ws = wb.create_sheet("Instructions", 0)

    ws.merge_cells("A1:D1")
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = CENTER
    ws.row_dimensions[1].height = 32

    ws["A3"] = "Purpose:"
    ws["A3"].font = SUB_HEADER_FONT
    ws.merge_cells("B3:D3")
    ws["B3"] = purpose
    ws["B3"].font = NORMAL_FONT
    ws["B3"].alignment = Alignment(wrap_text=True, vertical="top")

    ws["A5"] = "How to use:"
    ws["A5"].font = SUB_HEADER_FONT
    for i, step in enumerate(usage_steps, start=1):
        ws.merge_cells(start_row=5+i, start_column=2, end_row=5+i, end_column=4)
        ws.cell(row=5+i, column=1, value=f"  Step {i}").font = NORMAL_FONT
        c = ws.cell(row=5+i, column=2, value=step)
        c.font = NORMAL_FONT
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[5+i].height = 20

    next_row = 5 + len(usage_steps) + 2
    ws.cell(row=next_row, column=1, value="Columns:").font = SUB_HEADER_FONT

    # Column doc header
    next_row += 1
    headers = ["Column", "Type", "Required", "Description"]
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=next_row, column=j, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER

    for row in columns_doc:
        next_row += 1
        for j, val in enumerate(row, start=1):
            c = ws.cell(row=next_row, column=j, value=val)
            c.font = NORMAL_FONT
            c.alignment = LEFT if j == 4 else CENTER
            c.border = BORDER

    if notes:
        next_row += 2
        ws.cell(row=next_row, column=1, value="Notes:").font = SUB_HEADER_FONT
        for note in notes:
            next_row += 1
            ws.merge_cells(start_row=next_row, start_column=1, end_row=next_row, end_column=4)
            c = ws.cell(row=next_row, column=1, value=f"  • {note}")
            c.font = NOTE_FONT
            c.alignment = Alignment(wrap_text=True, vertical="top")

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 56


# ===========================================
# 1. GENERAL LEDGER (GL) Template
# ===========================================
def build_gl():
    wb = Workbook()
    wb.remove(wb.active)

    add_instructions_sheet(wb,
        title="GL — General Ledger Template",
        purpose="ใช้สำหรับ upload ข้อมูล GL Journal Entries จากระบบบัญชี เพื่อนำมาคำนวณ Cash Flow",
        usage_steps=[
            "Export GL transactions from your accounting system (SAP/Oracle/Xero/Manual)",
            "Paste data into 'GL_Data' sheet starting at row 4",
            "Ensure each entry has matching Debit + Credit (balanced)",
            "Save file with name format: GL_[YYYY-MM]_[Entity].xlsx (e.g., GL_2026-03_HMW.xlsx)",
            "Upload via app: Accounting → Upload GL → Select file",
        ],
        columns_doc=[
            ("Date", "Date", "Yes", "Transaction date (YYYY-MM-DD)"),
            ("JE_No", "Text", "Yes", "Journal entry number (e.g., JE-2026-0001)"),
            ("Account_Code", "Text", "Yes", "GL account code (e.g., 1010, 4100)"),
            ("Account_Name", "Text", "Yes", "Account description"),
            ("Description", "Text", "No", "Transaction description / memo"),
            ("Debit", "Number", "Yes*", "Debit amount (0 if credit) — *one of Debit/Credit required"),
            ("Credit", "Number", "Yes*", "Credit amount (0 if debit)"),
            ("Reference", "Text", "No", "Source document (Invoice no., bill no., etc.)"),
            ("Entity", "Text", "Yes", f"Company code: {ENTITIES_NOTE}"),
            ("Segment", "Text", "No", "Business segment ID (e.g., HONDA, F1, O1)"),
        ],
        notes=[
            "Total Debit MUST equal Total Credit per JE_No (double-entry rule)",
            "Use empty cells or 0 for non-applicable Debit/Credit",
            "Dates outside the period will be flagged but not rejected",
        ],
    )

    ws = wb.create_sheet("GL_Data")
    columns = ["Date", "JE_No", "Account_Code", "Account_Name", "Description",
               "Debit", "Credit", "Reference", "Entity", "Segment"]
    samples = [
        (datetime(2026,3,1), "JE-2026-0001", "1010", "เงินสด-ธนาคารกสิกร",
         "รับชำระจากลูกค้า Honda - INV-001", 1500000, 0, "INV-2026-001", "HMW", "HONDA"),
        (datetime(2026,3,1), "JE-2026-0001", "1130", "ลูกหนี้การค้า",
         "ตัดลูกหนี้ INV-001", 0, 1500000, "INV-2026-001", "HMW", "HONDA"),
        (datetime(2026,3,2), "JE-2026-0002", "5100", "เงินเดือนพนักงาน",
         "จ่ายเงินเดือน มี.ค.", 850000, 0, "PAY-202603", "HMW", "O1"),
        (datetime(2026,3,2), "JE-2026-0002", "1010", "เงินสด-ธนาคารกสิกร",
         "จ่ายเงินเดือน มี.ค.", 0, 850000, "PAY-202603", "HMW", "O1"),
        (datetime(2026,3,5), "JE-2026-0003", "5200", "ดอกเบี้ยจ่าย",
         "ดอกเบี้ยเงินกู้ระยะยาว", 125000, 0, "LN-2025-01", "HMW", "F1"),
        (datetime(2026,3,5), "JE-2026-0003", "1010", "เงินสด-ธนาคารกสิกร",
         "ดอกเบี้ยเงินกู้ระยะยาว", 0, 125000, "LN-2025-01", "HMW", "F1"),
    ]
    setup_sheet(ws, "GL — General Ledger Entries", columns, samples,
        notes=[
            f"Entities: {ENTITIES_NOTE}",
            f"Segments (income): {SEGMENT_NOTE_IN}",
            f"Segments (expense): {SEGMENT_NOTE_OUT}",
        ],
        col_widths=[12, 16, 12, 28, 35, 14, 14, 16, 10, 12])

    # Summary sheet with formulas
    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = "GL Upload Summary"
    ws2["A1"].font = TITLE_FONT
    ws2.merge_cells("A1:C1")
    ws2["A3"] = "Total Debit:"
    ws2["B3"] = "=SUM(GL_Data!F:F)"
    ws2["B3"].number_format = '#,##0.00'
    ws2["A4"] = "Total Credit:"
    ws2["B4"] = "=SUM(GL_Data!G:G)"
    ws2["B4"].number_format = '#,##0.00'
    ws2["A5"] = "Difference (must be 0):"
    ws2["B5"] = "=B3-B4"
    ws2["B5"].number_format = '#,##0.00'
    ws2["A6"] = "Total Entries:"
    ws2["B6"] = "=COUNTA(GL_Data!B:B)-3"
    for row in ["A3","A4","A5","A6"]:
        ws2[row].font = SUB_HEADER_FONT
    ws2.column_dimensions["A"].width = 24
    ws2.column_dimensions["B"].width = 18

    wb.save(os.path.join(OUT_DIR, "01_GL_Template.xlsx"))
    print("✓ 01_GL_Template.xlsx")


# ===========================================
# 2. TRIAL BALANCE Template
# ===========================================
def build_tb():
    wb = Workbook()
    wb.remove(wb.active)

    add_instructions_sheet(wb,
        title="Trial Balance Template",
        purpose="ใช้สำหรับ upload Trial Balance รายเดือน เพื่อใช้ใน Indirect Cash Flow calculation",
        usage_steps=[
            "Export Trial Balance from accounting system at month-end",
            "Paste data into 'TB_Data' sheet starting at row 4",
            "Verify Total Debit = Total Credit (Summary sheet will validate)",
            "Save file as TB_[YYYY-MM]_[Entity].xlsx",
            "Upload via app: Accounting → Upload Trial Balance",
        ],
        columns_doc=[
            ("Period", "Text", "Yes", "Period (YYYY-MM, e.g., 2026-03)"),
            ("Account_Code", "Text", "Yes", "GL account code"),
            ("Account_Name", "Text", "Yes", "Account name"),
            ("Account_Type", "Text", "Yes", "Asset/Liability/Equity/Revenue/Expense"),
            ("Opening_Debit", "Number", "No", "Opening balance (debit side)"),
            ("Opening_Credit", "Number", "No", "Opening balance (credit side)"),
            ("Period_Debit", "Number", "No", "Movements during period (debit)"),
            ("Period_Credit", "Number", "No", "Movements during period (credit)"),
            ("Closing_Debit", "Number", "Yes*", "Closing balance (debit)"),
            ("Closing_Credit", "Number", "Yes*", "Closing balance (credit)"),
            ("Entity", "Text", "Yes", f"Company: {ENTITIES_NOTE}"),
        ],
        notes=[
            "Account_Type drives Indirect Method calculations",
            "AR change, AP change, Inventory change are auto-derived from period movements",
        ],
    )

    ws = wb.create_sheet("TB_Data")
    columns = ["Period","Account_Code","Account_Name","Account_Type",
               "Opening_Debit","Opening_Credit","Period_Debit","Period_Credit",
               "Closing_Debit","Closing_Credit","Entity"]
    samples = [
        ("2026-03","1010","เงินสดและรายการเทียบเท่า","Asset",450000000,0,80000000,65000000,465000000,0,"HMW"),
        ("2026-03","1130","ลูกหนี้การค้า","Asset",125000000,0,45000000,38000000,132000000,0,"HMW"),
        ("2026-03","1200","สินค้าคงเหลือ","Asset",85000000,0,22000000,18000000,89000000,0,"HMW"),
        ("2026-03","1500","ที่ดิน อาคารและอุปกรณ์","Asset",520000000,0,15000000,5000000,530000000,0,"HMW"),
        ("2026-03","2110","เจ้าหนี้การค้า","Liability",0,95000000,30000000,42000000,0,107000000,"HMW"),
        ("2026-03","2210","เงินกู้ระยะสั้น","Liability",0,180000000,12000000,18000000,0,186000000,"HMW"),
        ("2026-03","3000","ทุนจดทะเบียน","Equity",0,500000000,0,0,0,500000000,"HMW"),
        ("2026-03","4100","รายได้จากการขาย","Revenue",0,0,0,85000000,0,85000000,"HMW"),
        ("2026-03","5100","ต้นทุนขาย","Expense",0,0,62000000,0,62000000,0,"HMW"),
        ("2026-03","5200","ค่าใช้จ่ายในการบริหาร","Expense",0,0,12500000,0,12500000,0,"HMW"),
    ]
    setup_sheet(ws, "Trial Balance Data", columns, samples,
        notes=[f"Entities: {ENTITIES_NOTE}",
               "Account_Type: Asset/Liability/Equity/Revenue/Expense"],
        col_widths=[10, 12, 30, 12, 14, 14, 14, 14, 14, 14, 10])

    # Summary with validation
    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = "Trial Balance Summary"
    ws2["A1"].font = TITLE_FONT
    ws2.merge_cells("A1:C1")
    ws2["A3"] = "Total Closing Debit:"
    ws2["B3"] = "=SUM(TB_Data!I:I)"
    ws2["A4"] = "Total Closing Credit:"
    ws2["B4"] = "=SUM(TB_Data!J:J)"
    ws2["A5"] = "Difference:"
    ws2["B5"] = "=B3-B4"
    ws2["A6"] = "Balanced?"
    ws2["B6"] = '=IF(ABS(B5)<0.01,"YES ✓","NO ✗")'
    for r in ["A3","A4","A5","A6"]:
        ws2[r].font = SUB_HEADER_FONT
    for r in ["B3","B4","B5"]:
        ws2[r].number_format = '#,##0.00'
    ws2.column_dimensions["A"].width = 24
    ws2.column_dimensions["B"].width = 20

    wb.save(os.path.join(OUT_DIR, "02_Trial_Balance_Template.xlsx"))
    print("✓ 02_Trial_Balance_Template.xlsx")


# ===========================================
# 3. AR / Invoices Template
# ===========================================
def build_ar():
    wb = Workbook()
    wb.remove(wb.active)

    add_instructions_sheet(wb,
        title="AR / Invoices Template",
        purpose="ใช้ upload ใบแจ้งหนี้ (Invoices) เพื่อวิเคราะห์ลูกหนี้และคาดการณ์เงินสดรับ",
        usage_steps=[
            "Export AR list from your accounting/billing system",
            "Paste into 'Invoices' sheet starting at row 4",
            "Ensure Due_Date is calculated from Date + Credit_Term_Days",
            "Status: Open / Paid / Partial / Overdue",
            "Upload via app: Finance → Upload AR/Invoices",
        ],
        columns_doc=[
            ("Invoice_No", "Text", "Yes", "Invoice number"),
            ("Date", "Date", "Yes", "Invoice date"),
            ("Customer_Code", "Text", "Yes", "Customer code (links to Customer_Master)"),
            ("Customer_Name", "Text", "Yes", "Customer name"),
            ("Customer_Type", "Text", "Yes", "Cash / Credit"),
            ("Amount", "Number", "Yes", "Invoice total amount"),
            ("Credit_Term_Days", "Number", "No", "Payment terms in days (0=cash)"),
            ("Due_Date", "Date", "Yes", "Payment due date (= Date + Credit_Term_Days)"),
            ("Amount_Paid", "Number", "No", "Amount received so far"),
            ("Balance", "Number", "Auto", "Outstanding = Amount - Amount_Paid"),
            ("Status", "Text", "Yes", "Open / Paid / Partial / Overdue"),
            ("Entity", "Text", "Yes", f"Receiving company: {ENTITIES_NOTE}"),
            ("Segment", "Text", "No", f"Revenue segment: {SEGMENT_NOTE_IN}"),
        ],
    )

    ws = wb.create_sheet("Invoices")
    columns = ["Invoice_No","Date","Customer_Code","Customer_Name","Customer_Type",
               "Amount","Credit_Term_Days","Due_Date","Amount_Paid","Balance",
               "Status","Entity","Segment"]
    samples = [
        ("INV-2026-0145", datetime(2026,3,1), "C001", "บริษัท ABC จำกัด", "Credit",
         1500000, 30, datetime(2026,3,31), 0, "=F5-I5", "Open", "HMW", "HONDA"),
        ("INV-2026-0146", datetime(2026,3,3), "C002", "ลูกค้าทั่วไป", "Cash",
         350000, 0, datetime(2026,3,3), 350000, "=F6-I6", "Paid", "HMW", "TOK"),
        ("INV-2026-0147", datetime(2026,2,15), "C003", "บริษัท XYZ Corp", "Credit",
         2200000, 60, datetime(2026,4,16), 1100000, "=F7-I7", "Partial", "HMW", "FIN"),
        ("INV-2026-0148", datetime(2026,1,10), "C004", "บริษัท Late Pay Ltd.", "Credit",
         800000, 30, datetime(2026,2,9), 0, "=F8-I8", "Overdue", "HMW", "HONDA"),
        ("INV-2026-0149", datetime(2026,3,10), "C005", "ACG Holding (Internal)", "Credit",
         500000, 30, datetime(2026,4,9), 0, "=F9-I9", "Open", "ACG", "MGT_HMW"),
    ]
    setup_sheet(ws, "AR / Invoices", columns, samples,
        notes=[
            "Balance column uses formula: Amount - Amount_Paid",
            f"Segments (income): {SEGMENT_NOTE_IN}",
        ],
        col_widths=[16,12,14,28,14,14,14,14,14,14,12,10,12])

    # Aging Summary with formulas
    ws2 = wb.create_sheet("Aging_Summary")
    ws2["A1"] = "AR Aging Summary"
    ws2["A1"].font = TITLE_FONT
    ws2.merge_cells("A1:E1")
    ws2["A3"] = "Bucket"
    ws2["B3"] = "Count"
    ws2["C3"] = "Total Outstanding"
    for c in ["A3","B3","C3"]:
        ws2[c].font = HEADER_FONT
        ws2[c].fill = HEADER_FILL
        ws2[c].alignment = CENTER

    ws2["A4"] = "Current (Not Due)"
    ws2["B4"] = '=COUNTIF(Invoices!K:K,"Open")'
    ws2["C4"] = '=SUMIF(Invoices!K:K,"Open",Invoices!J:J)'

    ws2["A5"] = "Partial"
    ws2["B5"] = '=COUNTIF(Invoices!K:K,"Partial")'
    ws2["C5"] = '=SUMIF(Invoices!K:K,"Partial",Invoices!J:J)'

    ws2["A6"] = "Overdue"
    ws2["B6"] = '=COUNTIF(Invoices!K:K,"Overdue")'
    ws2["C6"] = '=SUMIF(Invoices!K:K,"Overdue",Invoices!J:J)'

    ws2["A7"] = "Paid"
    ws2["B7"] = '=COUNTIF(Invoices!K:K,"Paid")'
    ws2["C7"] = '=SUMIF(Invoices!K:K,"Paid",Invoices!J:J)'

    ws2["A8"] = "TOTAL"
    ws2["A8"].font = SUB_HEADER_FONT
    ws2["B8"] = "=SUM(B4:B7)"
    ws2["C8"] = "=SUM(C4:C7)"
    ws2["C8"].font = SUB_HEADER_FONT

    for r in range(4, 9):
        ws2[f"C{r}"].number_format = '#,##0.00'

    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 12
    ws2.column_dimensions["C"].width = 22

    wb.save(os.path.join(OUT_DIR, "03_AR_Invoices_Template.xlsx"))
    print("✓ 03_AR_Invoices_Template.xlsx")


# ===========================================
# 4. AP / Bills Template
# ===========================================
def build_ap():
    wb = Workbook()
    wb.remove(wb.active)

    add_instructions_sheet(wb,
        title="AP / Bills Template",
        purpose="ใช้ upload ใบวางบิลจากผู้ขาย เพื่อคาดการณ์เงินสดจ่าย",
        usage_steps=[
            "Export AP list from accounting system",
            "Paste into 'Bills' sheet starting at row 4",
            "Status: Open / Paid / Partial / Overdue",
            "Upload via app: Finance → Upload AP/Bills",
        ],
        columns_doc=[
            ("Bill_No", "Text", "Yes", "Bill / Invoice number from vendor"),
            ("Date", "Date", "Yes", "Bill date"),
            ("Vendor_Code", "Text", "Yes", "Vendor code (links to Vendor_Master)"),
            ("Vendor_Name", "Text", "Yes", "Vendor name"),
            ("Category", "Text", "Yes", "Expense category"),
            ("Amount", "Number", "Yes", "Bill total amount"),
            ("Credit_Term_Days", "Number", "No", "Payment terms in days"),
            ("Due_Date", "Date", "Yes", "Payment due date"),
            ("Amount_Paid", "Number", "No", "Amount paid so far"),
            ("Balance", "Number", "Auto", "Outstanding"),
            ("Status", "Text", "Yes", "Open / Paid / Partial / Overdue"),
            ("Entity", "Text", "Yes", f"Paying company: {ENTITIES_NOTE}"),
            ("CF_Group", "Text", "No", f"Cash flow group: {SEGMENT_NOTE_OUT}"),
        ],
    )

    ws = wb.create_sheet("Bills")
    columns = ["Bill_No","Date","Vendor_Code","Vendor_Name","Category",
               "Amount","Credit_Term_Days","Due_Date","Amount_Paid","Balance",
               "Status","Entity","CF_Group"]
    samples = [
        ("BL-2026-0301", datetime(2026,3,1), "V001", "Honda Automobile (Thailand)", "Inventory",
         2500000, 45, datetime(2026,4,15), 0, "=F5-I5", "Open", "HMW", "O2"),
        ("BL-2026-0302", datetime(2026,3,2), "V002", "การไฟฟ้านครหลวง", "Utility",
         85000, 30, datetime(2026,4,1), 85000, "=F6-I6", "Paid", "HMW", "O4"),
        ("BL-2026-0303", datetime(2026,3,3), "V003", "Kasikorn Bank - Loan", "Interest",
         125000, 0, datetime(2026,3,3), 125000, "=F7-I7", "Paid", "HMW", "F1"),
        ("BL-2026-0304", datetime(2026,2,15), "V004", "บริษัท ก่อสร้าง CCC จำกัด", "Subcontract",
         580000, 30, datetime(2026,3,17), 0, "=F8-I8", "Overdue", "HMW", "O3"),
        ("BL-2026-0305", datetime(2026,3,10), "V005", "กรมสรรพากร", "Tax",
         245000, 30, datetime(2026,4,9), 0, "=F9-I9", "Open", "HMW", "O5"),
    ]
    setup_sheet(ws, "AP / Bills", columns, samples,
        notes=[f"CF_Group: {SEGMENT_NOTE_OUT}"],
        col_widths=[16,12,12,32,14,14,14,14,14,14,12,10,12])

    wb.save(os.path.join(OUT_DIR, "04_AP_Bills_Template.xlsx"))
    print("✓ 04_AP_Bills_Template.xlsx")


# ===========================================
# 5. Bank Transactions Template
# ===========================================
def build_bank():
    wb = Workbook()
    wb.remove(wb.active)

    add_instructions_sheet(wb,
        title="Bank Transactions Template",
        purpose="ใช้ upload รายการเดินบัญชีธนาคารเพื่อ Reconciliation กับ GL",
        usage_steps=[
            "Export bank statement (CSV/Excel) from your bank online portal",
            "Paste data into 'Bank_Transactions' sheet starting at row 4",
            "Type: IN (deposit/credit) or OUT (withdrawal/debit)",
            "Match Reference to invoice/bill numbers if possible",
            "Upload via app: Finance → Bank Transactions",
        ],
        columns_doc=[
            ("Date", "Date", "Yes", "Transaction date"),
            ("Bank_Account", "Text", "Yes", "Bank account name/number"),
            ("Type", "Text", "Yes", "IN (เงินเข้า) / OUT (เงินออก)"),
            ("Amount", "Number", "Yes", "Transaction amount (positive)"),
            ("Description", "Text", "Yes", "Bank statement description"),
            ("Reference", "Text", "No", "Match to invoice/bill number"),
            ("Category", "Text", "No", "Suggested category"),
            ("Balance", "Number", "No", "Running balance after transaction"),
            ("Entity", "Text", "Yes", f"Account holder: {ENTITIES_NOTE}"),
            ("Status", "Text", "No", "Matched / Unmatched / Manual"),
        ],
    )

    ws = wb.create_sheet("Bank_Transactions")
    columns = ["Date","Bank_Account","Type","Amount","Description","Reference",
               "Category","Balance","Entity","Status"]
    samples = [
        (datetime(2026,3,1), "Kasikorn-001-2-34567-8", "IN", 1500000,
         "TT INCOMING ABC CO", "INV-2026-0145", "Sales-Honda", 466500000, "HMW", "Matched"),
        (datetime(2026,3,2), "Kasikorn-001-2-34567-8", "OUT", 850000,
         "PAYROLL MAR2026", "PAY-202603", "Staff Salary", 465650000, "HMW", "Matched"),
        (datetime(2026,3,3), "Kasikorn-001-2-34567-8", "OUT", 125000,
         "LOAN INT PMT", "LN-2025-01", "Interest", 465525000, "HMW", "Matched"),
        (datetime(2026,3,4), "Kasikorn-001-2-34567-8", "IN", 350000,
         "CASH DEPOSIT", "INV-2026-0146", "Sales-Service", 465875000, "HMW", "Matched"),
        (datetime(2026,3,5), "Kasikorn-001-2-34567-8", "OUT", 85000,
         "BILL PAYMENT MEA", "BL-2026-0302", "Utility", 465790000, "HMW", "Matched"),
        (datetime(2026,3,6), "Kasikorn-001-2-34567-8", "OUT", 42000,
         "TRANSFER FEE", "", "Bank Fee", 465748000, "HMW", "Unmatched"),
    ]
    setup_sheet(ws, "Bank Transactions", columns, samples,
        col_widths=[12,22,8,14,28,16,16,16,10,12])

    # Reconciliation summary
    ws2 = wb.create_sheet("Reconciliation")
    ws2["A1"] = "Bank Reconciliation Summary"
    ws2["A1"].font = TITLE_FONT
    ws2.merge_cells("A1:C1")
    ws2["A3"] = "Total IN:"
    ws2["B3"] = '=SUMIF(Bank_Transactions!C:C,"IN",Bank_Transactions!D:D)'
    ws2["A4"] = "Total OUT:"
    ws2["B4"] = '=SUMIF(Bank_Transactions!C:C,"OUT",Bank_Transactions!D:D)'
    ws2["A5"] = "Net Movement:"
    ws2["B5"] = "=B3-B4"
    ws2["A7"] = "Matched:"
    ws2["B7"] = '=COUNTIF(Bank_Transactions!J:J,"Matched")'
    ws2["A8"] = "Unmatched:"
    ws2["B8"] = '=COUNTIF(Bank_Transactions!J:J,"Unmatched")'
    for r in ["A3","A4","A5","A7","A8"]:
        ws2[r].font = SUB_HEADER_FONT
    for r in ["B3","B4","B5"]:
        ws2[r].number_format = '#,##0.00'
    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 20

    wb.save(os.path.join(OUT_DIR, "05_Bank_Transactions_Template.xlsx"))
    print("✓ 05_Bank_Transactions_Template.xlsx")


# ===========================================
# 6. Cash Receipts Template
# ===========================================
def build_receipts():
    wb = Workbook()
    wb.remove(wb.active)

    add_instructions_sheet(wb,
        title="Cash Receipts Template",
        purpose="บันทึกการรับชำระเงินสดจากลูกค้า (ทั้งสดและเครดิต)",
        usage_steps=[
            "บันทึกการรับชำระแต่ละครั้ง",
            "Payment_Method: Cash / Transfer / Cheque / Credit Card",
            "Link Invoice_No เพื่อตัดยอดลูกหนี้",
            "Upload via app: Finance → Cash Receipts",
        ],
        columns_doc=[
            ("Date", "Date", "Yes", "Receipt date"),
            ("Receipt_No", "Text", "Yes", "Receipt number"),
            ("Customer_Code", "Text", "Yes", "Customer code"),
            ("Customer_Name", "Text", "Yes", "Customer name"),
            ("Invoice_No", "Text", "No", "Invoice being paid (for AR matching)"),
            ("Amount", "Number", "Yes", "Amount received"),
            ("Payment_Method", "Text", "Yes", "Cash / Transfer / Cheque / Credit Card"),
            ("Bank_Account", "Text", "No", "Receiving bank account"),
            ("Entity", "Text", "Yes", f"Receiving company: {ENTITIES_NOTE}"),
            ("Segment", "Text", "No", f"Revenue segment: {SEGMENT_NOTE_IN}"),
            ("Notes", "Text", "No", "Additional notes"),
        ],
    )

    ws = wb.create_sheet("Receipts")
    columns = ["Date","Receipt_No","Customer_Code","Customer_Name","Invoice_No",
               "Amount","Payment_Method","Bank_Account","Entity","Segment","Notes"]
    samples = [
        (datetime(2026,3,1), "RCP-2026-0201", "C001", "บริษัท ABC จำกัด", "INV-2026-0145",
         1500000, "Transfer", "Kasikorn-001-2-34567-8", "HMW", "HONDA", "ครบยอด"),
        (datetime(2026,3,3), "RCP-2026-0202", "C002", "ลูกค้าทั่วไป", "INV-2026-0146",
         350000, "Cash", "เงินสดมือ", "HMW", "TOK", ""),
        (datetime(2026,3,5), "RCP-2026-0203", "C003", "บริษัท XYZ Corp", "INV-2026-0147",
         1100000, "Cheque", "Kasikorn-001-2-34567-8", "HMW", "FIN", "ครึ่งยอด รอครึ่งหลัง"),
    ]
    setup_sheet(ws, "Cash Receipts", columns, samples,
        col_widths=[12,16,12,28,16,14,14,22,10,12,28])

    wb.save(os.path.join(OUT_DIR, "06_Cash_Receipts_Template.xlsx"))
    print("✓ 06_Cash_Receipts_Template.xlsx")


# ===========================================
# 7. Cash Payments Template
# ===========================================
def build_payments():
    wb = Workbook()
    wb.remove(wb.active)

    add_instructions_sheet(wb,
        title="Cash Payments Template",
        purpose="บันทึกการจ่ายเงินสดให้ผู้จัดจำหน่าย พนักงาน และค่าใช้จ่ายอื่นๆ",
        usage_steps=[
            "บันทึกการจ่ายเงินแต่ละรายการ",
            "Payment_Method: Cash / Transfer / Cheque",
            "Link Bill_No เพื่อตัดยอดเจ้าหนี้",
            "Upload via app: Finance → Cash Payments",
        ],
        columns_doc=[
            ("Date", "Date", "Yes", "Payment date"),
            ("Payment_No", "Text", "Yes", "Payment voucher number"),
            ("Vendor_Code", "Text", "Yes", "Vendor code"),
            ("Vendor_Name", "Text", "Yes", "Vendor name"),
            ("Bill_No", "Text", "No", "Bill being paid"),
            ("Amount", "Number", "Yes", "Payment amount"),
            ("Payment_Method", "Text", "Yes", "Cash / Transfer / Cheque"),
            ("Bank_Account", "Text", "No", "Paying bank account"),
            ("Entity", "Text", "Yes", f"Paying company: {ENTITIES_NOTE}"),
            ("CF_Group", "Text", "No", f"CF Group: {SEGMENT_NOTE_OUT}"),
            ("Notes", "Text", "No", "Additional notes"),
        ],
    )

    ws = wb.create_sheet("Payments")
    columns = ["Date","Payment_No","Vendor_Code","Vendor_Name","Bill_No",
               "Amount","Payment_Method","Bank_Account","Entity","CF_Group","Notes"]
    samples = [
        (datetime(2026,3,2), "PV-2026-0301", "V006", "พนักงาน (Payroll)", "PAY-202603",
         850000, "Transfer", "Kasikorn-001-2-34567-8", "HMW", "O1", "เงินเดือน มี.ค."),
        (datetime(2026,3,3), "PV-2026-0302", "V003", "Kasikorn Bank - Loan", "BL-2026-0303",
         125000, "Transfer", "Kasikorn-001-2-34567-8", "HMW", "F1", "ดอกเบี้ยเงินกู้"),
        (datetime(2026,3,4), "PV-2026-0303", "V002", "การไฟฟ้านครหลวง", "BL-2026-0302",
         85000, "Transfer", "Kasikorn-001-2-34567-8", "HMW", "O4", "ค่าไฟ มี.ค."),
        (datetime(2026,3,10), "PV-2026-0304", "V001", "Honda Automobile (Thailand)", "BL-2026-0301-partial",
         500000, "Cheque", "Kasikorn-001-2-34567-8", "HMW", "O2", "ชำระบางส่วน"),
    ]
    setup_sheet(ws, "Cash Payments", columns, samples,
        col_widths=[12,16,12,30,18,14,14,22,10,12,28])

    wb.save(os.path.join(OUT_DIR, "07_Cash_Payments_Template.xlsx"))
    print("✓ 07_Cash_Payments_Template.xlsx")


# ===========================================
# 8. Customer Master Template
# ===========================================
def build_customer_master():
    wb = Workbook()
    wb.remove(wb.active)

    add_instructions_sheet(wb,
        title="Customer Master Template",
        purpose="ข้อมูลหลักของลูกค้า รวม Credit Term เพื่อคาดการณ์ Cash collection",
        usage_steps=[
            "Maintain customer database with credit terms",
            "Customer_Type: Cash (เงินสดทุกครั้ง) / Credit (มีเครดิต)",
            "Credit_Term_Days: 0 for cash customers, 30/60/90 typically for credit",
            "Update once per customer, not per transaction",
        ],
        columns_doc=[
            ("Customer_Code", "Text", "Yes", "Unique customer code"),
            ("Customer_Name", "Text", "Yes", "Customer name"),
            ("Customer_Type", "Text", "Yes", "Cash / Credit"),
            ("Credit_Term_Days", "Number", "Yes", "0 for cash, days for credit"),
            ("Credit_Limit", "Number", "No", "Maximum credit allowed"),
            ("Contact_Person", "Text", "No", "Primary contact"),
            ("Phone", "Text", "No", "Contact phone"),
            ("Email", "Text", "No", "Contact email"),
            ("Address", "Text", "No", "Billing address"),
            ("Tax_ID", "Text", "No", "Tax identification (เลขประจำตัวผู้เสียภาษี)"),
            ("Entity", "Text", "Yes", f"Selling company: {ENTITIES_NOTE}"),
            ("Status", "Text", "Yes", "Active / Inactive"),
        ],
    )

    ws = wb.create_sheet("Customers")
    columns = ["Customer_Code","Customer_Name","Customer_Type","Credit_Term_Days",
               "Credit_Limit","Contact_Person","Phone","Email","Address","Tax_ID",
               "Entity","Status"]
    samples = [
        ("C001","บริษัท ABC จำกัด","Credit",30,5000000,"คุณสมชาย","02-123-4567",
         "abc@email.com","123 ถ.สุขุมวิท กรุงเทพ","0105560000123","HMW","Active"),
        ("C002","ลูกค้าทั่วไป","Cash",0,0,"","","","","","HMW","Active"),
        ("C003","บริษัท XYZ Corp","Credit",60,10000000,"Khun Smith","02-555-6666",
         "smith@xyz.com","99 อโศก กรุงเทพ","0105560000456","HMW","Active"),
        ("C004","บริษัท Late Pay Ltd.","Credit",30,1000000,"คุณวิชัย","02-777-8888",
         "wichai@latepay.com","45 รัชดา กรุงเทพ","0105560000789","HMW","Active"),
        ("C005","ACG Holding (Internal)","Credit",30,20000000,"Internal","Internal",
         "internal@acg.com","Internal Group","0105560000111","ACG","Active"),
    ]
    setup_sheet(ws, "Customer Master", columns, samples,
        col_widths=[12,28,12,14,16,16,14,22,30,16,10,10])

    wb.save(os.path.join(OUT_DIR, "08_Customer_Master_Template.xlsx"))
    print("✓ 08_Customer_Master_Template.xlsx")


# ===========================================
# 9. Vendor Master Template
# ===========================================
def build_vendor_master():
    wb = Workbook()
    wb.remove(wb.active)

    add_instructions_sheet(wb,
        title="Vendor Master Template",
        purpose="ข้อมูลหลักของผู้จัดจำหน่าย รวม Credit Term เพื่อคาดการณ์การจ่ายเงิน",
        usage_steps=[
            "Maintain vendor database with payment terms",
            "Category: Inventory / Utility / Subcontract / Tax / Bank / Other",
            "Update once per vendor, not per bill",
        ],
        columns_doc=[
            ("Vendor_Code", "Text", "Yes", "Unique vendor code"),
            ("Vendor_Name", "Text", "Yes", "Vendor name"),
            ("Category", "Text", "Yes", "Expense category"),
            ("Credit_Term_Days", "Number", "Yes", "Payment terms in days"),
            ("Contact_Person", "Text", "No", "Primary contact"),
            ("Phone", "Text", "No", "Contact phone"),
            ("Email", "Text", "No", "Contact email"),
            ("Address", "Text", "No", "Vendor address"),
            ("Tax_ID", "Text", "No", "Tax ID"),
            ("Bank_Account", "Text", "No", "Vendor's bank for payment"),
            ("Entity", "Text", "Yes", f"Paying company: {ENTITIES_NOTE}"),
            ("CF_Group", "Text", "No", f"CF Group: {SEGMENT_NOTE_OUT}"),
            ("Status", "Text", "Yes", "Active / Inactive"),
        ],
    )

    ws = wb.create_sheet("Vendors")
    columns = ["Vendor_Code","Vendor_Name","Category","Credit_Term_Days",
               "Contact_Person","Phone","Email","Address","Tax_ID","Bank_Account",
               "Entity","CF_Group","Status"]
    samples = [
        ("V001","Honda Automobile (Thailand)","Inventory",45,"Sales Dept","02-100-2000",
         "sales@honda.co.th","Honda HQ","0105560222000","BKK-Bank-111","HMW","O2","Active"),
        ("V002","การไฟฟ้านครหลวง","Utility",30,"Customer Service","1130",
         "info@mea.or.th","กรุงเทพ","0994000000999","MEA-Account","HMW","O4","Active"),
        ("V003","Kasikorn Bank - Loan","Interest",0,"Loan Officer","02-888-8888",
         "loan@kbank.com","Kasikorn HQ","0105560000888","Auto-debit","HMW","F1","Active"),
        ("V004","บริษัท ก่อสร้าง CCC จำกัด","Subcontract",30,"คุณประยุทธ์","02-444-5555",
         "ccc@build.com","นนทบุรี","0125560111222","SCB-555","HMW","O3","Active"),
        ("V005","กรมสรรพากร","Tax",30,"E-Filing","1161",
         "info@rd.go.th","กรุงเทพ","0994000000111","RD-Account","HMW","O5","Active"),
        ("V006","พนักงาน (Payroll)","Staff",0,"HR Dept","Internal",
         "hr@hmw.com","Internal","Internal","Various","HMW","O1","Active"),
    ]
    setup_sheet(ws, "Vendor Master", columns, samples,
        col_widths=[12,30,14,14,16,14,22,28,16,16,10,10,10])

    wb.save(os.path.join(OUT_DIR, "09_Vendor_Master_Template.xlsx"))
    print("✓ 09_Vendor_Master_Template.xlsx")


# ===========================================
# 10. Inventory Template
# ===========================================
def build_inventory():
    wb = Workbook()
    wb.remove(wb.active)

    add_instructions_sheet(wb,
        title="Inventory Template",
        purpose="ใช้ upload ยอดสินค้าคงเหลือสิ้นเดือน เพื่อคำนวณ Inventory Change ใน Indirect Method",
        usage_steps=[
            "Stock-take at month-end (or system snapshot)",
            "Period format: YYYY-MM (e.g., 2026-03)",
            "Total_Value = Quantity × Unit_Cost (auto-calculated)",
            "Upload via app: Accounting → Inventory Upload",
        ],
        columns_doc=[
            ("Period", "Text", "Yes", "Snapshot period (YYYY-MM)"),
            ("Item_Code", "Text", "Yes", "SKU / item code"),
            ("Item_Name", "Text", "Yes", "Item description"),
            ("Category", "Text", "No", "Item category"),
            ("Unit", "Text", "No", "Unit of measure (pcs, kg, etc.)"),
            ("Quantity", "Number", "Yes", "Quantity on hand"),
            ("Unit_Cost", "Number", "Yes", "Cost per unit"),
            ("Total_Value", "Number", "Auto", "= Quantity × Unit_Cost"),
            ("Location", "Text", "No", "Warehouse / store location"),
            ("Entity", "Text", "Yes", f"Owning company: {ENTITIES_NOTE}"),
        ],
    )

    ws = wb.create_sheet("Inventory")
    columns = ["Period","Item_Code","Item_Name","Category","Unit",
               "Quantity","Unit_Cost","Total_Value","Location","Entity"]
    samples = [
        ("2026-03","SKU-H001","Honda Wave 125i","Motorcycle","pcs",
         42,52000,"=F5*G5","คลังกลาง","HMW"),
        ("2026-03","SKU-H002","Honda Click 160i","Motorcycle","pcs",
         28,68000,"=F6*G6","คลังกลาง","HMW"),
        ("2026-03","SKU-H003","Honda PCX 160","Motorcycle","pcs",
         15,95000,"=F7*G7","คลังกลาง","HMW"),
        ("2026-03","SKU-SP01","อะไหล่ - น้ำมันเครื่อง","Parts","ขวด",
         350,180,"=F8*G8","คลัง Service","HMW"),
        ("2026-03","SKU-SP02","อะไหล่ - แบตเตอรี่","Parts","pcs",
         85,1850,"=F9*G9","คลัง Service","HMW"),
    ]
    setup_sheet(ws, "Inventory Snapshot", columns, samples,
        col_widths=[10,12,28,14,8,12,12,16,16,10])

    # Inventory change summary
    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = "Inventory Summary"
    ws2["A1"].font = TITLE_FONT
    ws2.merge_cells("A1:C1")
    ws2["A3"] = "Total Items:"
    ws2["B3"] = "=COUNTA(Inventory!B:B)-3"
    ws2["A4"] = "Total Value:"
    ws2["B4"] = "=SUM(Inventory!H:H)"
    ws2["B4"].number_format = '#,##0.00'
    for r in ["A3","A4"]:
        ws2[r].font = SUB_HEADER_FONT
    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 20

    wb.save(os.path.join(OUT_DIR, "10_Inventory_Template.xlsx"))
    print("✓ 10_Inventory_Template.xlsx")


# ===========================================
# 11. CF Mapping Template
# ===========================================
def build_cf_mapping():
    wb = Workbook()
    wb.remove(wb.active)

    add_instructions_sheet(wb,
        title="Cash Flow Mapping Template",
        purpose="กำหนดความสัมพันธ์ระหว่าง Segment → Activity → CF Line",
        usage_steps=[
            "List every segment and which CF report line it maps to",
            "Activity: Operating / Investing / Financing",
            "Type: Income / Expense",
            "Update once when setting up, then occasionally when adding segments",
            "Upload via app: Settings → Cash Flow Mapping",
        ],
        columns_doc=[
            ("Activity", "Text", "Yes", "Operating / Investing / Financing"),
            ("CF_Line", "Text", "Yes", "Thai cash flow line name"),
            ("Type", "Text", "Yes", "Income / Expense"),
            ("Group", "Text", "Yes", "INC (income) / F1 / I1 / O1-O6"),
            ("Entity", "Text", "Yes", f"{ENTITIES_NOTE} / *  (* = shared)"),
            ("Segment_ID", "Text", "Yes", "Segment code"),
            ("Segment_Name", "Text", "No", "Display name"),
        ],
    )

    ws = wb.create_sheet("CF_Mapping")
    columns = ["Activity","CF_Line","Type","Group","Entity","Segment_ID","Segment_Name"]
    samples = [
        # Operating - Income
        ("Operating","เงินสดรับจากการขายสินค้าและบริการ","Income","INC","HMW","HONDA","Honda (ขายรถ)"),
        ("Operating","เงินสดรับจากการขายสินค้าและบริการ","Income","INC","HMW","TOK","ตอกเข็ม"),
        ("Operating","เงินสดรับจากการขายสินค้าและบริการ","Income","INC","HMW","INS","ประกัน"),
        ("Operating","เงินสดรับจากการขายสินค้าและบริการ","Income","INC","HMW","FIN","ไฟแนนซ์"),
        ("Operating","เงินสดรับจากการขายสินค้าและบริการ","Income","INC","HMW","RENT","รถเช่า"),
        ("Operating","เงินสดรับจากการขายสินค้าและบริการ","Income","INC","HMW","DEL","รถส่ง"),
        ("Operating","เงินสดรับจากการขายสินค้าและบริการ","Income","INC","ACG","MGT_HMW","ค่าบริหาร HMW"),
        ("Operating","เงินสดรับจากการขายสินค้าและบริการ","Income","INC","ACG","MGT_CLIK","ค่าบริหาร CLIK"),
        ("Operating","เงินสดรับจากการขายสินค้าและบริการ","Income","INC","CLIK","SVC","Service"),
        # Operating - Expense
        ("Operating","เงินสดจ่ายให้แก่ผู้จัดจำหน่ายและพนักงาน","Expense","O1","*","O1","Staff exp"),
        ("Operating","เงินสดจ่ายให้แก่ผู้จัดจำหน่ายและพนักงาน","Expense","O2","*","O2","Inventory"),
        ("Operating","เงินสดจ่ายให้แก่ผู้จัดจำหน่ายและพนักงาน","Expense","O3","*","O3","Sub contract"),
        ("Operating","เงินสดจ่ายให้แก่ผู้จัดจำหน่ายและพนักงาน","Expense","O4","*","O4","Utility"),
        ("Operating","เงินสดจ่ายให้แก่ผู้จัดจำหน่ายและพนักงาน","Expense","O6","*","O6","Other"),
        ("Operating","เงินสดจ่ายดอกเบี้ย","Expense","F1","*","F1","Loan & Int"),
        ("Operating","เงินสดจ่ายภาษีเงินได้","Expense","O5","*","O5","Tax"),
        # Investing
        ("Investing","ซื้อที่ดิน อาคารและอุปกรณ์","Expense","I1","*","I1","Fixed assets"),
        # Financing
        ("Financing","เงินกู้ระยะสั้นเพิ่มขึ้น/(ลดลง)","Expense","F1","*","F1","Loan & Int"),
        ("Financing","เงินกู้ระยะยาว – ชำระคืน","Expense","F1","*","F1","Loan & Int"),
        ("Financing","จ่ายเงินปันผล","Expense","F1","*","F1","Loan & Int"),
    ]
    setup_sheet(ws, "Cash Flow Mapping", columns, samples,
        notes=[
            "* = shared across all entities (typical for expenses)",
            f"Income segments per entity: {SEGMENT_NOTE_IN}",
            f"Expense groups: {SEGMENT_NOTE_OUT}",
        ],
        col_widths=[14,42,12,10,10,14,22])

    wb.save(os.path.join(OUT_DIR, "11_CF_Mapping_Template.xlsx"))
    print("✓ 11_CF_Mapping_Template.xlsx")


# ===========================================
# 12. README / Index
# ===========================================
def build_readme():
    wb = Workbook()
    ws = wb.active
    ws.title = "Templates_Index"

    ws.merge_cells("A1:D1")
    ws["A1"] = "Cashflow Automation — Excel Templates Index"
    ws["A1"].font = Font(name=FONT_NAME, bold=True, size=16, color="2A6FF0")
    ws["A1"].alignment = CENTER
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:D2")
    ws["A2"] = "11 templates สำหรับ upload ข้อมูลเข้าระบบ Cashflow Automation"
    ws["A2"].font = NOTE_FONT
    ws["A2"].alignment = CENTER

    headers = ["#", "Template File", "Module", "Purpose"]
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=4, column=j, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = BORDER
    ws.row_dimensions[4].height = 28

    files = [
        (1, "01_GL_Template.xlsx", "Accounting",
         "General Ledger journal entries — main data source for everything"),
        (2, "02_Trial_Balance_Template.xlsx", "Accounting",
         "Monthly TB for Indirect Method calculations"),
        (3, "03_AR_Invoices_Template.xlsx", "AR / Finance",
         "Invoices to customers — feeds AR Aging + cash forecast"),
        (4, "04_AP_Bills_Template.xlsx", "AP / Finance",
         "Bills from vendors — feeds AP Aging + cash payment schedule"),
        (5, "05_Bank_Transactions_Template.xlsx", "Finance",
         "Bank statement for reconciliation"),
        (6, "06_Cash_Receipts_Template.xlsx", "Finance",
         "Cash collected (Direct Method input)"),
        (7, "07_Cash_Payments_Template.xlsx", "Finance",
         "Cash paid (Direct Method input)"),
        (8, "08_Customer_Master_Template.xlsx", "Master Data",
         "Customer database with credit terms"),
        (9, "09_Vendor_Master_Template.xlsx", "Master Data",
         "Vendor database with payment terms"),
        (10, "10_Inventory_Template.xlsx", "Accounting",
         "Month-end inventory snapshot"),
        (11, "11_CF_Mapping_Template.xlsx", "Settings",
         "Segment → CF Line mapping configuration"),
    ]
    for i, (n, fname, mod, desc) in enumerate(files):
        r = 5 + i
        ws.cell(row=r, column=1, value=n).alignment = CENTER
        ws.cell(row=r, column=2, value=fname).font = Font(name=FONT_NAME, color="2A6FF0", underline="single", size=10)
        ws.cell(row=r, column=3, value=mod).alignment = CENTER
        ws.cell(row=r, column=4, value=desc)
        for col in range(1, 5):
            ws.cell(row=r, column=col).border = BORDER
            if ws.cell(row=r, column=col).font.name is None:
                ws.cell(row=r, column=col).font = NORMAL_FONT

    # Upload workflow note
    next_r = 5 + len(files) + 2
    ws.cell(row=next_r, column=1, value="Upload Workflow:").font = SUB_HEADER_FONT
    workflow = [
        "1. Customer & Vendor Master — setup once (templates 8, 9)",
        "2. CF Mapping — setup once (template 11)",
        "3. GL or Trial Balance — upload monthly (templates 1, 2)",
        "4. AR, AP, Bank — upload weekly/monthly (templates 3, 4, 5)",
        "5. Receipts, Payments — upload daily/weekly if not in GL (6, 7)",
        "6. Inventory — upload at month-end (template 10)",
    ]
    for i, line in enumerate(workflow, start=1):
        ws.merge_cells(start_row=next_r+i, start_column=1, end_row=next_r+i, end_column=4)
        c = ws.cell(row=next_r+i, column=1, value=line)
        c.font = NORMAL_FONT
        c.alignment = LEFT

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 60

    ws.freeze_panes = "A5"

    wb.save(os.path.join(OUT_DIR, "00_README_Index.xlsx"))
    print("✓ 00_README_Index.xlsx")


# ===========================================
# Run all builders
# ===========================================
if __name__ == "__main__":
    print("Building Cashflow Automation Excel templates...")
    print("=" * 50)
    build_readme()
    build_gl()
    build_tb()
    build_ar()
    build_ap()
    build_bank()
    build_receipts()
    build_payments()
    build_customer_master()
    build_vendor_master()
    build_inventory()
    build_cf_mapping()
    print("=" * 50)
    print(f"\nAll templates created in: {OUT_DIR}")
    print(f"Total: 12 files")
