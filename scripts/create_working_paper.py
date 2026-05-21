#!/usr/bin/env python3
"""
Generate Working Paper Excel Export
6 sheets: AR Detail, AP Detail, Inv Detail, CF by Segment, Cash Reconciliation, Direct vs Indirect
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import json
import sys

def fmt(n):
    """Format number for display"""
    return f"{n:,.0f}"

def add_header_row(sheet, row_num, headers, bg_color="366092", text_color="FFFFFF"):
    """Add header row with formatting"""
    for col, header in enumerate(headers, 1):
        cell = sheet.cell(row_num, col, header)
        cell.font = Font(bold=True, size=11, color=text_color)
        cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.row_dimensions[row_num].height = 25

def ar_detail_sheet(wb, ar_data):
    """Create AR Detail sheet with 3 levels: Transaction → GL → Bank"""
    ws = wb.create_sheet("AR Detail", 0)
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 18

    row = 1
    ws.cell(row, 1, "AR Detail Reconciliation").font = Font(bold=True, size=14)
    ws.merge_cells(f'A{row}:E{row}')
    row += 1

    # Level 1: Transaction Summary
    add_header_row(ws, row, ["Entity", "Segment", "Opening", "Sales", "Collection", "Closing", "Change CF"], "4472C4")
    row += 1

    entities = list(ar_data.keys())
    for ent in entities:
        segs = ar_data[ent]['segments']
        total_opening = sum(s['opening'] for s in segs)
        total_sales = sum(s['sales'] for s in segs)
        total_collection = sum(s['collection'] for s in segs)
        total_closing = sum(s['closing'] for s in segs)
        total_change = total_closing - total_opening
        total_cf = -total_change

        # Entity summary row
        ws.cell(row, 1, ent).font = Font(bold=True)
        ws.cell(row, 3, total_opening).number_format = '#,##0'
        ws.cell(row, 4, total_sales).number_format = '#,##0'
        ws.cell(row, 5, total_collection).number_format = '#,##0'
        ws.cell(row, 6, total_closing).number_format = '#,##0'
        ws.cell(row, 7, total_cf).number_format = '#,##0'
        row += 1

        # Segment rows
        for seg in segs:
            ws.cell(row, 2, seg['name'])
            ws.cell(row, 3, seg['opening']).number_format = '#,##0'
            ws.cell(row, 4, seg['sales']).number_format = '#,##0'
            ws.cell(row, 5, seg['collection']).number_format = '#,##0'
            ws.cell(row, 6, seg['closing']).number_format = '#,##0'
            ws.cell(row, 7, -(seg['closing'] - seg['opening'])).number_format = '#,##0'
            row += 1

        row += 1

    return row

def ap_detail_sheet(wb, ap_data):
    """Create AP Detail sheet with 3 levels"""
    ws = wb.create_sheet("AP Detail", 1)
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 18

    row = 1
    ws.cell(row, 1, "AP Detail Reconciliation").font = Font(bold=True, size=14)
    ws.merge_cells(f'A{row}:E{row}')
    row += 1

    add_header_row(ws, row, ["Entity", "Segment", "Opening", "Purchases", "Payment", "Closing", "Change CF"], "70AD47")
    row += 1

    entities = list(ap_data.keys())
    for ent in entities:
        segs = ap_data[ent]['segments']
        total_opening = sum(s['opening'] for s in segs)
        total_purchases = sum(s['purchases'] for s in segs)
        total_payment = sum(s['payment'] for s in segs)
        total_closing = sum(s['closing'] for s in segs)
        total_change = total_closing - total_opening
        total_cf = total_change

        # Entity summary row
        ws.cell(row, 1, ent).font = Font(bold=True)
        ws.cell(row, 3, total_opening).number_format = '#,##0'
        ws.cell(row, 4, total_purchases).number_format = '#,##0'
        ws.cell(row, 5, total_payment).number_format = '#,##0'
        ws.cell(row, 6, total_closing).number_format = '#,##0'
        ws.cell(row, 7, total_cf).number_format = '#,##0'
        row += 1

        # Segment rows
        for seg in segs:
            ws.cell(row, 2, seg['name'])
            ws.cell(row, 3, seg['opening']).number_format = '#,##0'
            ws.cell(row, 4, seg['purchases']).number_format = '#,##0'
            ws.cell(row, 5, seg['payment']).number_format = '#,##0'
            ws.cell(row, 6, seg['closing']).number_format = '#,##0'
            ws.cell(row, 7, seg['closing'] - seg['opening']).number_format = '#,##0'
            row += 1

        row += 1

def inv_detail_sheet(wb, inv_opening, inv_closing):
    """Create Inventory Detail sheet"""
    ws = wb.create_sheet("Inv Detail", 2)
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20

    row = 1
    ws.cell(row, 1, "Inventory Detail Reconciliation").font = Font(bold=True, size=14)
    row += 2

    add_header_row(ws, row, ["Description", "Amount (THB)"], "4472C4")
    row += 1

    ws.cell(row, 1, "Opening Inventory (GL 1200, Q4 2025)")
    ws.cell(row, 2, inv_opening).number_format = '#,##0'
    row += 1

    ws.cell(row, 1, "Purchases")
    row += 1

    ws.cell(row, 1, "COGS")
    row += 1

    ws.cell(row, 1, "Closing Inventory (GL 1200, Q1 2026)")
    ws.cell(row, 2, inv_closing).number_format = '#,##0'
    row += 1

    inv_change = inv_closing - inv_opening
    inv_cf = -inv_change

    ws.cell(row, 1, "Change in Inventory").font = Font(bold=True)
    ws.cell(row, 2, inv_change).number_format = '#,##0'
    ws.cell(row, 2).font = Font(bold=True)
    row += 1

    ws.cell(row, 1, "CF Impact (Inventory Increase = Use Cash)").font = Font(bold=True)
    ws.cell(row, 2, inv_cf).number_format = '#,##0'
    ws.cell(row, 2).font = Font(bold=True)

def cf_summary_sheet(wb, ar_data, ap_data, inv_opening, inv_closing):
    """Create CF by Segment sheet"""
    ws = wb.create_sheet("CF by Segment", 3)
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18

    row = 1
    ws.cell(row, 1, "Cash Flow Impact by Working Capital Component").font = Font(bold=True, size=14)
    ws.merge_cells(f'A{row}:D{row}')
    row += 2

    add_header_row(ws, row, ["Component", "Opening", "Closing", "Change"], "70AD47")
    row += 1

    # AR Summary
    ar_opening = sum(sum(s['opening'] for s in v['segments']) for v in ar_data.values())
    ar_closing = sum(sum(s['closing'] for s in v['segments']) for v in ar_data.values())
    ar_change = ar_closing - ar_opening
    ar_cf = -ar_change

    ws.cell(row, 1, "Accounts Receivable (AR)").font = Font(bold=True)
    ws.cell(row, 2, ar_opening).number_format = '#,##0'
    ws.cell(row, 3, ar_closing).number_format = '#,##0'
    ws.cell(row, 4, ar_cf).number_format = '#,##0'
    ws.cell(row, 4).fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    row += 1

    # AP Summary
    ap_opening = sum(sum(s['opening'] for s in v['segments']) for v in ap_data.values())
    ap_closing = sum(sum(s['closing'] for s in v['segments']) for v in ap_data.values())
    ap_change = ap_closing - ap_opening
    ap_cf = ap_change

    ws.cell(row, 1, "Accounts Payable (AP)").font = Font(bold=True)
    ws.cell(row, 2, ap_opening).number_format = '#,##0'
    ws.cell(row, 3, ap_closing).number_format = '#,##0'
    ws.cell(row, 4, ap_cf).number_format = '#,##0'
    ws.cell(row, 4).fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    row += 1

    # Inventory Summary
    inv_change = inv_closing - inv_opening
    inv_cf = -inv_change

    ws.cell(row, 1, "Inventory (INV)").font = Font(bold=True)
    ws.cell(row, 2, inv_opening).number_format = '#,##0'
    ws.cell(row, 3, inv_closing).number_format = '#,##0'
    ws.cell(row, 4, inv_cf).number_format = '#,##0'
    ws.cell(row, 4).fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    row += 2

    # Total WC Impact
    total_wc_cf = ar_cf + ap_cf + inv_cf
    ws.cell(row, 1, "Total Working Capital Impact").font = Font(bold=True, size=12)
    ws.cell(row, 4, total_wc_cf).number_format = '#,##0'
    ws.cell(row, 4).font = Font(bold=True, size=12)
    ws.cell(row, 4).fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")

def cash_reconciliation_sheet(wb, opening_cash=265_000_000):
    """Create Cash Reconciliation sheet"""
    ws = wb.create_sheet("Cash Reconciliation", 4)
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 20

    row = 1
    ws.cell(row, 1, "Cash Reconciliation (Direct Method)").font = Font(bold=True, size=14)
    row += 2

    add_header_row(ws, row, ["Description", "Amount (THB)"], "4472C4")
    row += 1

    ws.cell(row, 1, "Opening Cash (Q4 2025)")
    ws.cell(row, 2, opening_cash).number_format = '#,##0'
    row += 2

    ws.cell(row, 1, "Operating Activities").font = Font(bold=True)
    row += 1

    ws.cell(row, 1, "  Cash from Sales")
    row += 1

    ws.cell(row, 1, "  Less: AR Increase (use of cash)")
    row += 1

    ws.cell(row, 1, "  Less: Operating Expenses")
    row += 1

    ws.cell(row, 1, "  Plus: AP Increase (source of cash)")
    row += 2

    ws.cell(row, 1, "Investing Activities").font = Font(bold=True)
    row += 1

    ws.cell(row, 1, "  Less: Fixed Asset Purchases")
    row += 2

    ws.cell(row, 1, "Financing Activities").font = Font(bold=True)
    row += 1

    ws.cell(row, 1, "  Debt Proceeds")
    row += 1

    ws.cell(row, 1, "  Less: Debt Repayment")
    row += 2

    ws.cell(row, 1, "Closing Cash (Q1 2026)").font = Font(bold=True, size=12)
    row += 1

def direct_indirect_sheet(wb):
    """Create Direct vs Indirect summary sheet"""
    ws = wb.create_sheet("Direct vs Indirect", 5)
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20

    row = 1
    ws.cell(row, 1, "Direct vs Indirect Method Comparison").font = Font(bold=True, size=14)
    ws.merge_cells(f'A{row}:C{row}')
    row += 2

    add_header_row(ws, row, ["Description", "Direct Method", "Indirect Method"], "70AD47")
    row += 1

    # Operating CF
    ws.cell(row, 1, "Operating Cash Flow").font = Font(bold=True)
    ws.cell(row, 2, "=").number_format = '@'
    ws.cell(row, 3, "=").number_format = '@'
    row += 2

    ws.cell(row, 1, "Investing Cash Flow").font = Font(bold=True)
    row += 1

    ws.cell(row, 1, "Financing Cash Flow").font = Font(bold=True)
    row += 2

    ws.cell(row, 1, "Net Change in Cash").font = Font(bold=True, size=12)
    ws.cell(row, 2, "").font = Font(bold=True, size=12)
    ws.cell(row, 3, "").font = Font(bold=True, size=12)

    row += 2
    ws.cell(row, 1, "✓ Both methods should show same Net CF")
    ws.merge_cells(f'A{row}:C{row}')

def main():
    """Main execution"""
    # Sample AR data (from data.js reconcileARBySegment)
    ar_data = {
        "HMW": {
            "segments": [
                {"id": "HONDA", "name": "HONDA (ขายรถ)", "opening": 135_000_000, "sales": 248_000_000, "collection": -228_500_000, "closing": 154_500_000},
                {"id": "TOK", "name": "ตอกเข็ม", "opening": 11_000_000, "sales": 20_000_000, "collection": -18_500_000, "closing": 12_500_000},
                {"id": "INS", "name": "ประกัน", "opening": 15_000_000, "sales": 28_000_000, "collection": -26_000_000, "closing": 17_000_000},
                {"id": "INS_OPEN", "name": "ประกันภัยเปิด", "opening": 8_000_000, "sales": 14_000_000, "collection": -13_500_000, "closing": 8_500_000},
                {"id": "INS_LIM", "name": "ประวางวง", "opening": 5_000_000, "sales": 10_000_000, "collection": -9_800_000, "closing": 5_200_000},
                {"id": "FIN", "name": "ไฟแนนซ์", "opening": 21_000_000, "sales": 38_000_000, "collection": -37_500_000, "closing": 21_500_000},
                {"id": "RENT", "name": "รถเช่า", "opening": 11_000_000, "sales": 20_000_000, "collection": -19_300_000, "closing": 11_700_000},
                {"id": "DEL", "name": "รถส่ง", "opening": 7_000_000, "sales": 12_000_000, "collection": -12_100_000, "closing": 6_900_000},
                {"id": "OTH", "name": "อื่นๆ", "opening": 11_000_000, "sales": 22_400_000, "collection": -20_000_000, "closing": 13_400_000},
            ]
        },
        "CLIK": {
            "segments": [
                {"id": "SVC", "name": "Service", "opening": 78_000_000, "sales": 155_000_000, "collection": -140_000_000, "closing": 91_900_000},
                {"id": "OTH", "name": "อื่นๆ", "opening": 8_100_000, "sales": 10_300_000, "collection": -8_900_000, "closing": 9_400_000},
            ]
        },
        "ACG": {
            "segments": [
                {"id": "MGT_HMW", "name": "ค่าบริหารจัดการ HMW", "opening": 9_500_000, "sales": 22_000_000, "collection": -18_000_000, "closing": 13_500_000},
                {"id": "MGT_CLIK", "name": "ค่าบริหารจัดการ CLIK", "opening": 4_000_000, "sales": 9_000_000, "collection": -7_800_000, "closing": 5_200_000},
                {"id": "OTH", "name": "อื่นๆ", "opening": 1_000_000, "sales": 1_400_000, "collection": -1_300_000, "closing": 1_100_000},
            ]
        },
    }

    # Sample AP data (from data.js reconcileAPBySegment)
    ap_data = {
        "HMW": {
            "segments": [
                {"id": "F1", "name": "Loan & Int", "opening": 6_900_000, "purchases": 18_800_000, "payment": -17_200_000, "closing": 8_500_000},
                {"id": "I1", "name": "Fixed assets", "opening": 4_900_000, "purchases": 13_500_000, "payment": -12_300_000, "closing": 6_100_000},
                {"id": "O1", "name": "Staff exp", "opening": 16_700_000, "purchases": 45_700_000, "payment": -41_700_000, "closing": 20_700_000},
                {"id": "O2", "name": "Inventory", "opening": 51_200_000, "purchases": 139_800_000, "payment": -127_600_000, "closing": 63_400_000},
                {"id": "O3", "name": "Sub contract", "opening": 3_900_000, "purchases": 10_800_000, "payment": -9_800_000, "closing": 4_900_000},
                {"id": "O4", "name": "Rental & Fac.", "opening": 4_900_000, "purchases": 13_400_000, "payment": -12_200_000, "closing": 6_100_000},
                {"id": "O5", "name": "Tax", "opening": 3_000_000, "purchases": 8_100_000, "payment": -7_300_000, "closing": 3_800_000},
                {"id": "O6", "name": "Others", "opening": 6_900_000, "purchases": 18_800_000, "payment": -17_200_000, "closing": 8_500_000},
            ]
        },
        "CLIK": {
            "segments": [
                {"id": "F1", "name": "Loan & Int", "opening": 1_900_000, "purchases": 5_200_000, "payment": -4_800_000, "closing": 2_300_000},
                {"id": "O1", "name": "Staff exp", "opening": 8_900_000, "purchases": 24_300_000, "payment": -22_100_000, "closing": 11_100_000},
                {"id": "O2", "name": "Inventory", "opening": 15_600_000, "purchases": 42_800_000, "payment": -38_900_000, "closing": 19_500_000},
                {"id": "O3", "name": "Sub contract", "opening": 2_300_000, "purchases": 6_200_000, "payment": -5_700_000, "closing": 2_800_000},
                {"id": "O4", "name": "Rental & Fac.", "opening": 2_100_000, "purchases": 5_800_000, "payment": -5_200_000, "closing": 2_700_000},
                {"id": "O6", "name": "Others", "opening": 1_500_000, "purchases": 4_100_000, "payment": -3_700_000, "closing": 1_900_000},
            ]
        },
        "ACG": {
            "segments": [
                {"id": "F1", "name": "Loan & Int", "opening": 800_000, "purchases": 2_200_000, "payment": -2_000_000, "closing": 1_000_000},
                {"id": "O1", "name": "Staff exp", "opening": 3_200_000, "purchases": 8_800_000, "payment": -8_000_000, "closing": 4_000_000},
                {"id": "O6", "name": "Others", "opening": 600_000, "purchases": 1_600_000, "payment": -1_500_000, "closing": 700_000},
            ]
        },
    }

    # Inventory data (from GL account 1200)
    inv_opening = 342_500_000  # Q4 2025
    inv_closing = 365_200_000  # Q1 2026

    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Create all sheets
    ar_detail_sheet(wb, ar_data)
    ap_detail_sheet(wb, ap_data)
    inv_detail_sheet(wb, inv_opening, inv_closing)
    cf_summary_sheet(wb, ar_data, ap_data, inv_opening, inv_closing)
    cash_reconciliation_sheet(wb)
    direct_indirect_sheet(wb)

    # Save workbook
    output_path = "Working_Paper_WC_Drill_Down.xlsx"
    wb.save(output_path)
    print(f"[OK] Created {output_path}")
    print(f"     6 sheets: AR Detail, AP Detail, Inv Detail, CF by Segment, Cash Reconciliation, Direct vs Indirect")
    return output_path

if __name__ == "__main__":
    main()
