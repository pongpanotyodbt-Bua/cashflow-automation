import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

# ── CF mapping by account code prefix ──────────────────────────────────────
def get_cf_mapping(code, bs_class):
    p4 = code[:4]
    p3 = code[:3]
    p2 = code[:2]

    # ── BALANCE SHEET ITEMS ──
    if bs_class in ("Asset", "Liabilities", "Equity"):

        # Cash & Bank
        if p4 in ("1111", "1112", "1113"):
            return ("Operating", "Cash & Bank", "Direct: Opening/Closing Cash")

        # Trade Receivables
        if p4 in ("1121", "1122", "1123", "1124", "1125", "1126", "1127", "1128", "1129"):
            return ("Operating", "Trade Receivables (AR)", "Indirect WC: AR change")

        # Other Current Assets
        if p3 in ("113", "114", "115", "116", "117", "118"):
            return ("Operating", "Other Current Assets", "Indirect WC: Other CA change")

        # Inventories
        if p4.startswith("12"):
            return ("Operating", "Inventories", "Indirect WC: Inventory change")

        # Long-term investments
        if p4.startswith("13"):
            return ("Investing", "Long-term Investment", "Investing: Purchase/Sale of investments")

        # PPE / Fixed Assets
        if p4.startswith("14") or p4.startswith("15"):
            return ("Investing", "PPE & Intangibles", "Investing: Capex / Disposal")

        # Trade Payables
        if p4 in ("2111", "2112", "2113", "2114", "2115", "2116", "2117", "2118", "2119"):
            return ("Operating", "Trade Payables (AP)", "Indirect WC: AP change")

        # Accrued expenses & other current liabilities
        if p3 in ("212", "213", "214", "215", "216", "217", "218", "219"):
            return ("Operating", "Accrued Liabilities", "Indirect WC: Other CL change")

        # Short-term debt
        if p4 in ("2211", "2212", "2213"):
            return ("Financing", "Short-term Borrowing", "Financing: Drawdown / Repayment")

        # Long-term debt
        if p4.startswith("23") or p4.startswith("24"):
            return ("Financing", "Long-term Debt", "Financing: Drawdown / Repayment")

        # Equity / Share capital
        if p4.startswith("31"):
            return ("Financing", "Share Capital", "Financing: Capital raise")

        # Retained earnings
        if p4.startswith("32") or p4 in ("3160", "3150"):
            return ("Financing", "Retained Earnings", "Residual: Net profit + dividends")

        return ("Memo", "Other BS Item", "")

    # ── PROFIT & LOSS ITEMS ──
    if bs_class == "PL":

        # Revenue - Sales / ขาย
        if p4.startswith("411") or p4.startswith("412") or p4.startswith("413") or p4.startswith("414") or p4.startswith("415"):
            return ("Operating", "Revenue from Sales", "Direct: Cash receipts from customers")

        # Revenue - Other income / รายได้อื่น
        if p4.startswith("421"):
            return ("Operating", "Other Income", "Direct: Other operating receipts")

        # COGS - Cost of Goods/Services / ต้นทุน
        if p4.startswith("511") or p4.startswith("512") or p4.startswith("513") or p4.startswith("514"):
            return ("Operating", "Cost of Sales (COGS)", "Direct: Cash paid to suppliers")

        # Selling & Distribution / ต้นทุนการจัดจำหน่าย
        if p4.startswith("611"):
            return ("Operating", "Selling Expense", "Direct: Cash paid - selling costs")

        # G&A / SG&A / ค่าใช้จ่ายบริหาร
        if p4.startswith("612") or p4.startswith("613") or p4.startswith("614"):
            return ("Operating", "G&A Expense", "Direct: Cash paid - admin costs")

        # Depreciation & Amortization
        if "depr" in code.lower() or p4 in ("6121",):
            return ("Operating", "Depreciation (non-cash)", "Indirect: Add-back depreciation")

        # Finance cost / ดอกเบี้ย
        if p4.startswith("711") or p4.startswith("712"):
            return ("Financing", "Finance Cost / Interest", "Financing: Interest paid")

        # Other income/expense below EBIT
        if p4.startswith("811") or p4.startswith("812"):
            return ("Operating", "Other Income/Expense", "Direct: Other receipts/payments")

        # Tax / ภาษี
        if p4.startswith("900"):
            return ("Operating", "Income Tax", "Direct: Tax paid")

        return ("Operating", "Other P&L", "Review for reclassification")

    return ("Memo", "Unclassified", "")


# ── Read extracted COA CSV ──────────────────────────────────────────────────
import csv

with open("data/COA_from_TB.csv", encoding="utf-8-sig") as f:
    coa_rows = list(csv.DictReader(f))

# ── Build Excel ─────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()

# ── Colors & Styles ──
DARK_NAVY  = "1A2744"
BLUE_HDR   = "2A6FF0"
GOLD_HDR   = "C09000"
GREEN_HDR  = "1F9D55"
RED_HDR    = "C0392B"
PURPLE_HDR = "7C4DFF"
TEAL_HDR   = "0E97A0"

LIGHT_BLUE   = "E8F0FE"
LIGHT_GREEN  = "E8F5E9"
LIGHT_RED    = "FDE8E8"
LIGHT_YELLOW = "FFF9E6"
LIGHT_PURPLE = "F3EFFE"
LIGHT_GRAY   = "F5F7FA"

WHITE = "FFFFFF"

CF_COLORS = {
    "Operating":  ("E8F5E9", "1F9D55"),
    "Investing":  ("E8F0FE", "2A6FF0"),
    "Financing":  ("FDE8E8", "C0392B"),
    "Memo":       ("F5F7FA", "8B95A1"),
}

def hdr_style(cell, bg_hex, bold=True, white_text=True, sz=10, center=True):
    cell.font = Font(bold=bold, color=WHITE if white_text else "000000", size=sz, name="Arial")
    cell.fill = PatternFill("solid", fgColor=bg_hex)
    cell.alignment = Alignment(horizontal="center" if center else "left",
                                vertical="center", wrap_text=True)

def border_thin():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def num_fmt(cell, val, fmt="#,##0.00;(#,##0.00);-"):
    if val == "" or val is None:
        cell.value = 0
    else:
        try:
            cell.value = float(val)
        except:
            cell.value = 0
    cell.number_format = fmt
    cell.font = Font(name="Arial", size=9)
    cell.alignment = Alignment(horizontal="right")

# ════════════════════════════════════════════════════════════════════════════
# SHEET 1 – Full COA with CF Mapping
# ════════════════════════════════════════════════════════════════════════════
ws = wb.active
ws.title = "COA & CF Mapping"
ws.freeze_panes = "A3"
ws.sheet_view.showGridLines = False

# Title row
ws.merge_cells("A1:S1")
ws["A1"] = "Chart of Accounts — CF Category Mapping  |  ACG / HMW / CLIK Group  |  2 Periods: Dec 2025 & Mar 2026"
ws["A1"].font = Font(bold=True, size=13, color=WHITE, name="Arial")
ws["A1"].fill = PatternFill("solid", fgColor=DARK_NAVY)
ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 30

# Column headers
headers = [
    ("A", 12, "Account Code"),
    ("B", 40, "Account Name (TH)"),
    ("C", 28, "Sub Caption (TH)"),
    ("D", 28, "FS Caption (TH)"),
    ("E", 12, "BS Class"),
    ("F", 14, "CF Activity"),
    ("G", 26, "CF Category"),
    ("H", 36, "CF Treatment Note"),
    ("I", 14, "Companies"),
    ("J", 15, "ACG Dec 2025"),
    ("K", 15, "HMW Dec 2025"),
    ("L", 15, "CLIK Dec 2025"),
    ("M", 15, "CONSO Dec 2025"),
    ("N", 15, "ACG Mar 2026"),
    ("O", 15, "HMW Mar 2026"),
    ("P", 15, "CLIK Mar 2026"),
    ("Q", 15, "CONSO Mar 2026"),
    ("R", 15, "Movement (Conso)"),
    ("S", 20, "CF Sign Note"),
]

col_hdr_colors = {
    "A": DARK_NAVY, "B": DARK_NAVY, "C": DARK_NAVY, "D": DARK_NAVY, "E": DARK_NAVY,
    "F": BLUE_HDR,  "G": BLUE_HDR,  "H": BLUE_HDR,
    "I": DARK_NAVY,
    "J": GOLD_HDR,  "K": GOLD_HDR,  "L": GOLD_HDR,  "M": GOLD_HDR,
    "N": GREEN_HDR, "O": GREEN_HDR, "P": GREEN_HDR, "Q": GREEN_HDR,
    "R": TEAL_HDR,
    "S": PURPLE_HDR,
}

for col_letter, col_width, col_name in headers:
    cell = ws[f"{col_letter}2"]
    cell.value = col_name
    hdr_style(cell, col_hdr_colors.get(col_letter, DARK_NAVY), sz=9)
    cell.border = border_thin()
    ws.column_dimensions[col_letter].width = col_width

ws.row_dimensions[2].height = 36

# Data rows
for i, row in enumerate(coa_rows, start=3):
    code = row["account_code"]
    bs_class = row["bs_class"]

    cf_activity, cf_category, cf_note = get_cf_mapping(code, bs_class)
    bg_fill, txt_color = CF_COLORS.get(cf_activity, ("FFFFFF", "000000"))

    # Alternating row tint
    row_bg = bg_fill if (i % 2 == 0) else "FFFFFF"

    def cell_val(col, value, is_num=False):
        c = ws.cell(row=i, column=col, value=value)
        if is_num:
            try:
                c.value = float(value) if value not in ("", None, "0") else 0
            except:
                c.value = 0
            c.number_format = "#,##0.00;(#,##0.00);-"
            c.alignment = Alignment(horizontal="right", vertical="center")
        else:
            c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
        c.font = Font(name="Arial", size=9, color="000000")
        c.fill = PatternFill("solid", fgColor=row_bg)
        c.border = border_thin()
        return c

    cell_val(1,  code)
    cell_val(2,  row["account_name_th"])
    cell_val(3,  row["sub_caption_th"])
    cell_val(4,  row["fs_caption_th"])

    c_class = cell_val(5, bs_class)
    c_class.alignment = Alignment(horizontal="center", vertical="center")

    c_act = cell_val(6, cf_activity)
    c_act.font = Font(name="Arial", size=9, bold=True, color=txt_color)
    c_act.alignment = Alignment(horizontal="center", vertical="center")

    cell_val(7, cf_category)
    cell_val(8, cf_note)

    c_co = cell_val(9, row["companies_active"])
    c_co.alignment = Alignment(horizontal="center", vertical="center")

    cell_val(10, row["acg_dec25"],  True)
    cell_val(11, row["hmw_dec25"],  True)
    cell_val(12, row["clik_dec25"], True)
    cell_val(13, row["conso_dec25"],True)
    cell_val(14, row["acg_mar26"],  True)
    cell_val(15, row["hmw_mar26"],  True)
    cell_val(16, row["clik_mar26"], True)
    cell_val(17, row["conso_mar26"],True)

    # Movement
    try:
        dec = float(row["conso_dec25"]) if row["conso_dec25"] not in ("", None) else 0
        mar = float(row["conso_mar26"]) if row["conso_mar26"] not in ("", None) else 0
        movement = mar - dec
    except:
        movement = 0
    c_mv = ws.cell(row=i, column=18, value=movement)
    c_mv.number_format = "+#,##0.00;(#,##0.00);-"
    c_mv.font = Font(name="Arial", size=9,
                     color=("1F9D55" if movement >= 0 else "C0392B"))
    c_mv.alignment = Alignment(horizontal="right", vertical="center")
    c_mv.fill = PatternFill("solid", fgColor=row_bg)
    c_mv.border = border_thin()

    # CF Sign note
    sign_notes = {
        "Cash & Bank": "↑ = Cash increase",
        "Trade Receivables (AR)": "↑ AR = Cash used (subtract)",
        "Trade Payables (AP)": "↑ AP = Cash source (add)",
        "Inventories": "↑ Inv = Cash used (subtract)",
        "Revenue from Sales": "Credit = Cash received",
        "Cost of Sales (COGS)": "Debit = Cash paid",
        "Finance Cost / Interest": "Debit = Interest paid",
        "PPE & Intangibles": "↑ = Capex (cash out)",
    }
    cell_val(19, sign_notes.get(cf_category, ""))

# Auto-filter
ws.auto_filter.ref = f"A2:S{len(coa_rows)+2}"

# ════════════════════════════════════════════════════════════════════════════
# SHEET 2 – Summary by CF Category
# ════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("CF Category Summary")
ws2.sheet_view.showGridLines = False

ws2.merge_cells("A1:H1")
ws2["A1"] = "CF Category Summary — Balance Movement Dec 2025 → Mar 2026"
ws2["A1"].font = Font(bold=True, size=13, color=WHITE, name="Arial")
ws2["A1"].fill = PatternFill("solid", fgColor=DARK_NAVY)
ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 28

sum_headers = ["CF Activity", "CF Category", "# Accounts", "Conso Dec 2025", "Conso Mar 2026",
               "Movement", "% Change", "CF Relevance Note"]
sum_widths  = [14, 32, 12, 18, 18, 18, 12, 40]

for j, (h, w) in enumerate(zip(sum_headers, sum_widths), 1):
    c = ws2.cell(row=2, column=j, value=h)
    hdr_style(c, DARK_NAVY, sz=9)
    c.border = border_thin()
    ws2.column_dimensions[get_column_letter(j)].width = w
ws2.row_dimensions[2].height = 32

from collections import defaultdict
summary = defaultdict(lambda: {"count": 0, "dec25": 0.0, "mar26": 0.0})
for row in coa_rows:
    code = row["account_code"]
    bs_class = row["bs_class"]
    cf_activity, cf_category, _ = get_cf_mapping(code, bs_class)
    key = (cf_activity, cf_category)
    summary[key]["count"] += 1
    try:
        summary[key]["dec25"] += float(row["conso_dec25"] or 0)
        summary[key]["mar26"] += float(row["conso_mar26"] or 0)
    except:
        pass

CF_RELEVANCE = {
    "Cash & Bank": "Main cash balance — reconcile to bank",
    "Trade Receivables (AR)": "WC: AR change affects Operating CF",
    "Trade Payables (AP)": "WC: AP change affects Operating CF",
    "Inventories": "WC: Inventory change affects Operating CF",
    "Other Current Assets": "WC: Review for prepaid, deposits",
    "Accrued Liabilities": "WC: Accruals affect Operating CF",
    "Revenue from Sales": "Direct CF: Cash received from customers",
    "Cost of Sales (COGS)": "Direct CF: Cash paid to suppliers",
    "Selling Expense": "Direct CF: Selling cash outflows",
    "G&A Expense": "Direct CF: Admin cash outflows",
    "Finance Cost / Interest": "Financing CF: Interest paid",
    "Income Tax": "Operating CF: Tax paid",
    "PPE & Intangibles": "Investing CF: Capex, disposals",
    "Long-term Investment": "Investing CF: Investment flows",
    "Short-term Borrowing": "Financing CF: ST loan movements",
    "Long-term Debt": "Financing CF: LT loan movements",
    "Share Capital": "Financing CF: Capital raise/buyback",
    "Retained Earnings": "Memo: Net income + dividends",
    "Depreciation (non-cash)": "Indirect: Non-cash add-back",
    "Other Income/Expense": "Review: classify per cash impact",
}

sorted_keys = sorted(summary.keys(), key=lambda x: (
    {"Operating": 0, "Investing": 1, "Financing": 2, "Memo": 3}.get(x[0], 9), x[1]))

activity_order = {"Operating": 0, "Investing": 1, "Financing": 2, "Memo": 3}
prev_activity = None

for r_idx, key in enumerate(sorted_keys, start=3):
    act, cat = key
    v = summary[key]
    dec = v["dec25"]
    mar = v["mar26"]
    mov = mar - dec
    pct = (mov / dec * 100) if dec != 0 else 0

    bg, tc = CF_COLORS.get(act, ("FFFFFF", "000000"))

    vals = [act, cat, v["count"], dec, mar, mov, pct / 100,
            CF_RELEVANCE.get(cat, "")]

    for j, val in enumerate(vals, 1):
        c = ws2.cell(row=r_idx, column=j, value=val)
        c.fill = PatternFill("solid", fgColor=bg if r_idx % 2 == 0 else "FFFFFF")
        c.border = border_thin()
        c.font = Font(name="Arial", size=9,
                      bold=(j in (1, 2)),
                      color=tc if j == 1 else "000000")
        c.alignment = Alignment(
            horizontal="center" if j in (1, 3) else ("right" if j in (4, 5, 6, 7) else "left"),
            vertical="center")
        if j in (4, 5, 6):
            c.number_format = "#,##0;(#,##0);-"
        elif j == 7:
            c.number_format = "+0.0%;(0.0%);-"
        if j == 6:
            c.font = Font(name="Arial", size=9, bold=True,
                          color=("1F9D55" if mov >= 0 else "C0392B"))
    ws2.row_dimensions[r_idx].height = 18

ws2.auto_filter.ref = f"A2:H{len(sorted_keys)+2}"
ws2.freeze_panes = "A3"

# ════════════════════════════════════════════════════════════════════════════
# SHEET 3 – CF Framework Template (Indirect Method)
# ════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("CF Framework (Indirect)")
ws3.sheet_view.showGridLines = False

INDIRECT_TEMPLATE = [
    # (level, label, cf_category, note)
    (0, "OPERATING ACTIVITIES", None, ""),
    (1, "Net Profit / (Loss)", "Retained Earnings", "From P&L / Equity"),
    (0, "  Adjustments for non-cash items:", None, ""),
    (1, "  + Depreciation & Amortization", "Depreciation (non-cash)", "Add-back; account 61xx, 51xx depr"),
    (1, "  + Loss on asset disposal", "Other P&L", "Add-back non-cash"),
    (1, "  ± Other non-cash adjustments", "Other P&L", "e.g. allowances, provisions"),
    (0, "  Changes in Working Capital:", None, ""),
    (1, "  ± Trade Receivables (AR)", "Trade Receivables (AR)", "↑AR = subtract; ↓AR = add"),
    (1, "  ± Inventories", "Inventories", "↑Inv = subtract; ↓Inv = add"),
    (1, "  ± Other Current Assets", "Other Current Assets", "↑OCA = subtract"),
    (1, "  ± Trade Payables (AP)", "Trade Payables (AP)", "↑AP = add; ↓AP = subtract"),
    (1, "  ± Accrued Liabilities", "Accrued Liabilities", "↑Accr = add"),
    (0, "  Cash Generated from Operations", None, "Subtotal before tax & interest"),
    (1, "  − Income Tax Paid", "Income Tax", ""),
    (1, "  − Interest Paid (if classified operating)", "Finance Cost / Interest", ""),
    (0, "NET CASH FROM OPERATING ACTIVITIES", None, "═══"),
    (0, "", None, ""),
    (0, "INVESTING ACTIVITIES", None, ""),
    (1, "  − Capex (Purchase of PPE)", "PPE & Intangibles", "Cash used for fixed asset purchase"),
    (1, "  + Proceeds from asset disposal", "PPE & Intangibles", ""),
    (1, "  ± Long-term investments", "Long-term Investment", ""),
    (0, "NET CASH FROM INVESTING ACTIVITIES", None, "═══"),
    (0, "", None, ""),
    (0, "FINANCING ACTIVITIES", None, ""),
    (1, "  + New borrowings (ST)", "Short-term Borrowing", "Drawdown"),
    (1, "  − Loan repayments (ST)", "Short-term Borrowing", "Repayment"),
    (1, "  + New borrowings (LT)", "Long-term Debt", "Drawdown"),
    (1, "  − Loan repayments (LT)", "Long-term Debt", "Repayment"),
    (1, "  + Capital injection", "Share Capital", ""),
    (1, "  − Dividends paid", "Retained Earnings", ""),
    (0, "NET CASH FROM FINANCING ACTIVITIES", None, "═══"),
    (0, "", None, ""),
    (0, "NET INCREASE / (DECREASE) IN CASH", None, "Sum of Operating + Investing + Financing"),
    (0, "+ Opening Cash Balance (Dec 2025)", "Cash & Bank", ""),
    (0, "= CLOSING CASH BALANCE (Mar 2026)", "Cash & Bank", "Must reconcile to TB 111xxxxx"),
]

ws3.merge_cells("A1:E1")
ws3["A1"] = "Cash Flow Statement — Indirect Method Framework  |  Q1 2026 (Jan–Mar 2026)"
ws3["A1"].font = Font(bold=True, size=13, color=WHITE, name="Arial")
ws3["A1"].fill = PatternFill("solid", fgColor=DARK_NAVY)
ws3["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws3.row_dimensions[1].height = 28

ws3_hdrs = ["Line Item", "CF Category Reference", "Amount (THB)", "Source TB Account(s)", "Note"]
ws3_widths = [48, 28, 20, 30, 44]
for j, (h, w) in enumerate(zip(ws3_hdrs, ws3_widths), 1):
    c = ws3.cell(row=2, column=j, value=h)
    hdr_style(c, BLUE_HDR, sz=9)
    c.border = border_thin()
    ws3.column_dimensions[get_column_letter(j)].width = w
ws3.row_dimensions[2].height = 28

for r_idx, (level, label, cat, note) in enumerate(INDIRECT_TEMPLATE, start=3):
    is_section = (level == 0 and label and not label.startswith(" "))
    is_blank = (label == "")

    c1 = ws3.cell(row=r_idx, column=1, value=label)
    c2 = ws3.cell(row=r_idx, column=2, value=cat or "")
    c3 = ws3.cell(row=r_idx, column=3, value=0 if not is_blank and not is_section else None)
    c4 = ws3.cell(row=r_idx, column=4, value="")
    c5 = ws3.cell(row=r_idx, column=5, value=note)

    if is_section:
        for c in [c1, c2, c3, c4, c5]:
            c.font = Font(name="Arial", size=10, bold=True, color=WHITE)
            c.fill = PatternFill("solid", fgColor=BLUE_HDR)
            c.border = border_thin()
        ws3.row_dimensions[r_idx].height = 22
    elif is_blank:
        for c in [c1, c2, c3, c4, c5]:
            c.fill = PatternFill("solid", fgColor="F5F7FA")
        ws3.row_dimensions[r_idx].height = 8
    else:
        bg = CF_COLORS.get("Operating", ("E8F5E9", "1F9D55"))[0]
        if cat in ("PPE & Intangibles", "Long-term Investment"):
            bg = CF_COLORS["Investing"][0]
        elif cat in ("Short-term Borrowing", "Long-term Debt", "Share Capital"):
            bg = CF_COLORS["Financing"][0]

        for c in [c1, c2, c3, c4, c5]:
            c.font = Font(name="Arial", size=9)
            c.fill = PatternFill("solid", fgColor=bg)
            c.border = border_thin()
        if c3.value == 0:
            c3.number_format = "#,##0;(#,##0);-"
            c3.alignment = Alignment(horizontal="right")
        ws3.row_dimensions[r_idx].height = 18

ws3.freeze_panes = "A3"

# ── Save ────────────────────────────────────────────────────────────────────
out_path = "data/COA_CF_Mapping.xlsx"
wb.save(out_path)
print(f"Saved: {out_path}")
print(f"Sheets: {wb.sheetnames}")
print(f"Total accounts mapped: {len(coa_rows)}")
