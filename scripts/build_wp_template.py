import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def fix(s):
    if isinstance(s, str):
        try:
            return s.encode('latin1').decode('cp874')
        except:
            return s
    return s

def num(x):
    return x if isinstance(x, (int, float)) else 0

# ── Load TB ──────────────────────────────────────────────────────────────────
tb = openpyxl.load_workbook('sample-data/TB by Entity&Conso.xlsx', read_only=True, data_only=True)
ws_tb = tb['Sheet1']
tb_rows = []
for i, row in enumerate(ws_tb.iter_rows(values_only=True)):
    if i == 0:
        continue
    if row[1] is None and row[3] is None:
        continue
    tb_rows.append(row)

# ── Section definitions ───────────────────────────────────────────────────────
SECTIONS = [
    {
        'title': '1. ลูกหนี้การค้า (Accounts Receivable)',
        'fs_kw': ['ลูกหนี้การค้า'],
        'cat': 'Asset',
        'cf_sign': -1,
        'cf_note': 'AR เพิ่ม = ใช้เงินสด = CF ลบ  |  AR ลด = เก็บเงินได้ = CF บวก',
        'activity': 'Operating',
    },
    {
        'title': '2. ลูกหนี้หมุนเวียนอื่น (Other Current Receivables)',
        'fs_kw': ['ลูกหนี้หมุนเวียนอื่น', 'สินทรัพย์ภาษีเงินได้ของงวดปัจจุบัน'],
        'cat': 'Asset',
        'cf_sign': -1,
        'cf_note': 'เพิ่ม = CF ลบ',
        'activity': 'Operating',
    },
    {
        'title': '3. ที่ดิน อาคาร อุปกรณ์ / PPE & Investment Property',
        'fs_kw': ['ที่ดิน อาคาร', 'อสังหาริมทรัพย์'],
        'cat': 'Asset',
        'cf_sign': -1,
        'cf_note': 'ดูร่วมกับรายการซื้อ/ขาย FA + ค่าเสื่อมที่ add-back ใน Operating',
        'activity': 'Investing',
    },
    {
        'title': '4. สินทรัพย์สิทธิการใช้ ROU Asset (TFRS 16)',
        'fs_kw': ['สินทรัพย์สิทธิ', 'สิทธิการใช้'],
        'cat': 'Asset',
        'cf_sign': 0,
        'cf_note': 'Non-cash — ค่าเสื่อม ROU บวกกลับใน Operating  |  ชำระ Lease เงินต้น = Financing',
        'activity': 'Operating (add-back)',
    },
    {
        'title': '5. สินทรัพย์ภาษีเงินได้รอตัดบัญชี / สินทรัพย์อื่น',
        'fs_kw': ['สินทรัพย์ภาษีเงินได้รอ', 'สินทรัพย์ทางการเงิน', 'เงินให้กู้ยืมระยะยาว', 'เงินลงทุน'],
        'cat': 'Asset',
        'cf_sign': -1,
        'cf_note': 'เพิ่ม = CF ลบ',
        'activity': 'Investing',
    },
    {
        'title': '6. เจ้าหนี้การค้า (Accounts Payable)',
        'fs_kw': ['เจ้าหนี้การค้า'],
        'cat': 'Liabilities',
        'cf_sign': 1,
        'cf_note': 'AP เพิ่ม = ยังไม่จ่าย = CF บวก  |  AP ลด = จ่ายแล้ว = CF ลบ',
        'activity': 'Operating',
    },
    {
        'title': '7. เจ้าหนี้หมุนเวียนอื่น / ค่าใช้จ่ายค้างจ่าย / ดอกเบี้ยค้างจ่าย',
        'fs_kw': ['เจ้าหนี้หมุนเวียนอื่น'],
        'cat': 'Liabilities',
        'cf_sign': 1,
        'cf_note': 'เพิ่ม = CF บวก  —  รวม: ค่าใช้จ่ายค้างจ่าย, ดอกเบี้ยค้างจ่าย, เงินรับล่วงหน้า',
        'activity': 'Operating',
    },
    {
        'title': '8. หนี้สินตามสัญญาเช่า Lease Liability (TFRS 16)',
        'fs_kw': ['หนี้สินตามสัญญาเช่า', 'ที่ถึงกำหนดชำระภายในหนึ่งปี'],
        'cat': 'Liabilities',
        'cf_sign': 0,
        'cf_note': 'เงินต้น Lease = Financing Activity  |  ดอกเบี้ย Lease = Operating หรือ Financing ตาม Policy',
        'activity': 'Financing',
    },
    {
        'title': '9. เงินกู้ยืมจากสถาบันการเงิน (Bank Loans)',
        'fs_kw': ['เงินกู้ยืมระยะสั้นจากสถาบัน', 'เงินเบิกเกินบัญชี', 'เงินกู้ยืมระยะยาวจากสถาบัน'],
        'cat': 'Liabilities',
        'cf_sign': 1,
        'cf_note': 'กู้เพิ่ม = CF บวก  |  ชำระคืน = CF ลบ — Financing Activity',
        'activity': 'Financing',
    },
    {
        'title': '10. ประมาณการผลประโยชน์พนักงาน (Employee Benefit)',
        'fs_kw': ['ผลประโยชน์พนักงาน', 'ประมาณการหนี้สิน'],
        'cat': 'Liabilities',
        'cf_sign': 1,
        'cf_note': 'Non-cash accrual — บวกกลับใน Operating',
        'activity': 'Operating (add-back)',
    },
    {
        'title': '11. ภาษีเงินได้นิติบุคคลค้างจ่าย (Income Tax Payable)',
        'fs_kw': ['ภาษีเงินได้นิติบุคคลค้างจ่าย'],
        'cat': 'Liabilities',
        'cf_sign': 1,
        'cf_note': 'เพิ่ม = CF บวก',
        'activity': 'Operating',
    },
    {
        'title': '12. หนี้สินภาษีเงินได้รอตัดบัญชี / หนี้สินไม่หมุนเวียนอื่น',
        'fs_kw': ['หนี้สินไม่หมุนเวียนอื่น'],
        'cat': 'Liabilities',
        'cf_sign': 1,
        'cf_note': 'เพิ่ม = CF บวก',
        'activity': 'Operating / Financing',
    },
]

IC_KEYWORDS = ['บริษัทย่อย', 'กิจการที่เกี่ยวข้อง', 'Share Service']

section_data = {s['title']: [] for s in SECTIONS}
unmatched_bs = []

for r in tb_rows:
    cat = str(r[0]) if r[0] else ''
    if cat in ('PL', 'None', 'Equity'):
        continue
    acc_no = str(r[2]) if r[2] else ''
    if not acc_no or acc_no in ('None', 'Check TB', 'Profit/Loss'):
        continue
    acc_name = fix(r[3]) if r[3] else ''
    fs_raw = r[5]
    fs = fix(fs_raw) if isinstance(fs_raw, str) else (str(fs_raw) if fs_raw else '')
    conso = num(r[19])
    acc_name_str = str(acc_name)
    is_ic = any(kw in acc_name_str for kw in IC_KEYWORDS)
    if not isinstance(fs, str): fs = ''
    if 'เงินสดและรายการเทียบเท่า' in fs:
        continue
    matched = False
    for sec in SECTIONS:
        exp_cat = sec.get('cat', '')
        if exp_cat and exp_cat != cat:
            continue
        if any(kw in fs for kw in sec['fs_kw']):
            section_data[sec['title']].append({
                'acc_no': acc_no, 'acc_name': acc_name,
                'fs': fs, 'conso': conso, 'is_ic': is_ic, 'cat': cat
            })
            matched = True
            break
    if not matched:
        unmatched_bs.append((cat, acc_no, acc_name, fs, conso))

# ── Build Excel ───────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()

# ── Styles ────────────────────────────────────────────────────────────────────
FONT_NAME = 'Arial'

def font(bold=False, color='000000', size=10, italic=False):
    return Font(name=FONT_NAME, bold=bold, color=color, size=size, italic=italic)

def fill(hex_color):
    return PatternFill('solid', start_color=hex_color, fgColor=hex_color)

def border_thin():
    s = Side(style='thin', color='CCCCCC')
    return Border(left=s, right=s, top=s, bottom=s)

def border_medium_bottom():
    return Border(bottom=Side(style='medium', color='999999'))

def align(h='left', v='center', wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

NUM_FMT = '#,##0.00_);(#,##0.00);"-"'
NUM_FMT_CF = '+#,##0.00;(#,##0.00);"-"'

# Colors
COL_HDR_BG   = '1F3864'   # dark navy
COL_HDR_FG   = 'FFFFFF'
COL_SEC_BG   = '2E75B6'   # blue section
COL_SEC_FG   = 'FFFFFF'
COL_SUB_BG   = 'D6E4F0'   # light blue sub-header
COL_IC_BG    = 'FFF2CC'   # yellow = IC
COL_TOTAL_BG = 'E2EFDA'   # light green total
COL_INPUT_FG = '0000FF'   # blue = input
COL_CALC_FG  = '000000'   # black = formula
COL_NOTE_FG  = '7F7F7F'   # gray = note
COL_WARN_BG  = 'FFE699'   # amber = intercompany warning

# ── Sheet 1: Working Paper ────────────────────────────────────────────────────
ws = wb.active
ws.title = 'Working Paper'

# Freeze panes
ws.freeze_panes = 'D4'

# Column widths
col_widths = {
    'A': 8,   # #
    'B': 14,  # Account No.
    'C': 48,  # Account Name
    'D': 18,  # FS Caption
    'E': 16,  # Opening Balance (INPUT)
    'F': 16,  # Closing Balance (from TB)
    'G': 16,  # Change
    'H': 16,  # CF Impact
    'I': 12,  # Activity
    'J': 8,   # IC Flag
    'K': 40,  # Note
}
for col_letter, width in col_widths.items():
    ws.column_dimensions[col_letter].width = width

# Row 1: Title
ws.merge_cells('A1:K1')
ws['A1'] = 'WORKING PAPER — Working Capital & Balance Sheet Movement'
ws['A1'].font = font(bold=True, size=13, color=COL_HDR_FG)
ws['A1'].fill = fill(COL_HDR_BG)
ws['A1'].alignment = align('center')
ws.row_dimensions[1].height = 22

# Row 2: Subtitle
ws.merge_cells('A2:K2')
ws['A2'] = 'Period: [กรอกงวด]     Company: Consolidated (CONSO)     Prepared by: ___________     Date: ___________'
ws['A2'].font = font(size=9, color='444444', italic=True)
ws['A2'].fill = fill('F2F2F2')
ws['A2'].alignment = align('left')
ws.row_dimensions[2].height = 16

# Row 3: Column headers
headers = ['#', 'Account No.', 'Account Name', 'FS Caption',
           'Opening Balance\n(กรอกต้นงวด)', 'Closing Balance\n(จาก TB)',
           'Change\n(Closing-Opening)', 'CF Impact\n(ผลต่อ CF)',
           'CF Activity', 'IC?', 'Note / Reference']
for col_idx, h in enumerate(headers, start=1):
    cell = ws.cell(row=3, column=col_idx, value=h)
    cell.font = font(bold=True, color=COL_HDR_FG, size=9)
    cell.fill = fill(COL_HDR_BG)
    cell.alignment = align('center', wrap=True)
    cell.border = border_thin()
ws.row_dimensions[3].height = 32

# ── Write sections ─────────────────────────────────────────────────────────────
current_row = 4
section_summary_rows = []   # (section_title, total_row)

for sec in SECTIONS:
    title = sec['title']
    accounts = section_data[title]
    cf_sign = sec['cf_sign']
    activity = sec['activity']
    cf_note = sec['cf_note']

    # Section header row
    ws.merge_cells(f'A{current_row}:K{current_row}')
    cell = ws.cell(row=current_row, column=1, value=title)
    cell.font = font(bold=True, color=COL_SEC_FG, size=10)
    cell.fill = fill(COL_SEC_BG)
    cell.alignment = align('left')
    ws.row_dimensions[current_row].height = 18
    sec_hdr_row = current_row
    current_row += 1

    # CF Note row
    ws.merge_cells(f'A{current_row}:K{current_row}')
    cell = ws.cell(row=current_row, column=1, value=f'   Logic: {cf_note}   |   Activity: {activity}')
    cell.font = font(size=8, color=COL_NOTE_FG, italic=True)
    cell.fill = fill('EEF4FB')
    cell.alignment = align('left')
    ws.row_dimensions[current_row].height = 14
    current_row += 1

    # Sub-header
    sub_hdrs = ['', '', 'Intercompany accounts highlighted in yellow', '', '', '', '', '', '', '', '']
    for col_idx, h in enumerate(sub_hdrs, start=1):
        cell = ws.cell(row=current_row, column=col_idx, value=h)
        cell.font = font(size=8, color='0000FF', italic=True)
        cell.fill = fill(COL_SUB_BG)
        cell.border = border_thin()
    ws.row_dimensions[current_row].height = 13
    current_row += 1

    data_start = current_row
    row_nums = []

    for acct in accounts:
        acc_no = acct['acc_no']
        acc_name = acct['acc_name']
        conso = acct['conso']
        is_ic = acct['is_ic']
        bg = COL_IC_BG if is_ic else 'FFFFFF'

        cols = [
            ('A', ''),
            ('B', acc_no),
            ('C', acc_name),
            ('D', acct['fs']),
            ('E', None),          # Opening — INPUT (blue)
            ('F', conso),         # Closing from TB
            ('G', f'=F{current_row}-E{current_row}'),   # Change
            ('H', None),          # CF Impact formula — set after
            ('I', activity),
            ('J', 'IC' if is_ic else ''),
            ('K', ''),
        ]

        for col_letter, val in cols:
            col_idx = ord(col_letter) - ord('A') + 1
            cell = ws.cell(row=current_row, column=col_idx, value=val)
            cell.fill = fill(bg)
            cell.border = border_thin()
            cell.alignment = align('left', 'center')

            if col_letter in ('E', 'F', 'G', 'H'):
                cell.alignment = align('right', 'center')
                cell.number_format = NUM_FMT

            if col_letter == 'E':
                cell.font = font(color=COL_INPUT_FG, size=9)
                cell.fill = fill('EAF3FF' if not is_ic else 'FFF5CC')
            elif col_letter == 'F':
                cell.font = font(color='006100', size=9)  # green = from TB
            elif col_letter == 'G':
                cell.font = font(color=COL_CALC_FG, size=9)
            elif col_letter == 'H':
                # CF Impact = Change * sign (for Assets: negative sign; for Liabilities: positive)
                if cf_sign == -1:
                    cell.value = f'=-G{current_row}'
                elif cf_sign == 1:
                    cell.value = f'=G{current_row}'
                else:
                    cell.value = 'N/A — see note'
                    cell.font = font(color='7F7F7F', size=9, italic=True)
                cell.number_format = NUM_FMT_CF if cf_sign != 0 else '@'
            elif col_letter == 'J' and is_ic:
                cell.font = font(color='7B3F00', bold=True, size=9)
                cell.fill = fill(COL_WARN_BG)
            else:
                cell.font = font(size=9)

        ws.row_dimensions[current_row].height = 15
        row_nums.append(current_row)
        current_row += 1

    # Total row for this section
    if row_nums:
        e_range = f'E{data_start}:E{current_row-1}'
        f_range = f'F{data_start}:F{current_row-1}'
        g_range = f'G{data_start}:G{current_row-1}'

        totals = [
            ('A', ''),
            ('B', ''),
            ('C', f'รวม {title}'),
            ('D', ''),
            ('E', f'=SUM(E{data_start}:E{current_row-1})'),
            ('F', f'=SUM(F{data_start}:F{current_row-1})'),
            ('G', f'=SUM(G{data_start}:G{current_row-1})'),
            ('H', f'=SUM(H{data_start}:H{current_row-1})' if cf_sign != 0 else 'N/A'),
            ('I', activity),
            ('J', ''),
            ('K', ''),
        ]
        for col_letter, val in totals:
            col_idx = ord(col_letter) - ord('A') + 1
            cell = ws.cell(row=current_row, column=col_idx, value=val)
            cell.font = font(bold=True, size=9, color='000000')
            cell.fill = fill(COL_TOTAL_BG)
            cell.border = border_thin()
            cell.alignment = align('left', 'center')
            if col_letter in ('E', 'F', 'G', 'H'):
                cell.alignment = align('right', 'center')
                cell.number_format = NUM_FMT_CF if col_letter == 'H' and cf_sign != 0 else NUM_FMT
        ws.row_dimensions[current_row].height = 16
        section_summary_rows.append((title, current_row, cf_sign, activity))
        current_row += 1

    # Spacer
    current_row += 1

# ── Sheet 2: CF Summary ───────────────────────────────────────────────────────
ws2 = wb.create_sheet('CF Summary (WC)')
ws2.column_dimensions['A'].width = 55
ws2.column_dimensions['B'].width = 18
ws2.column_dimensions['C'].width = 20

# Header
ws2.merge_cells('A1:C1')
ws2['A1'] = 'สรุปผลกระทบ Working Capital ต่อ Cash Flow Statement (Indirect Method)'
ws2['A1'].font = font(bold=True, size=12, color=COL_HDR_FG)
ws2['A1'].fill = fill(COL_HDR_BG)
ws2['A1'].alignment = align('center')
ws2.row_dimensions[1].height = 22

ws2['A2'] = 'รายการ'
ws2['B2'] = 'CF Impact (บาท)'
ws2['C2'] = 'Activity'
for col in ['A', 'B', 'C']:
    ws2[f'{col}2'].font = font(bold=True, color=COL_HDR_FG, size=10)
    ws2[f'{col}2'].fill = fill(COL_HDR_BG)
    ws2[f'{col}2'].alignment = align('center')
ws2.row_dimensions[2].height = 18

r = 3
op_rows = []
fin_rows = []
inv_rows = []

for (title, total_row, cf_sign, activity) in section_summary_rows:
    cell_a = ws2.cell(row=r, column=1, value=title)
    cell_b = ws2.cell(row=r, column=2)
    cell_c = ws2.cell(row=r, column=3, value=activity)

    if cf_sign != 0:
        cell_b.value = f"='Working Paper'!H{total_row}"
        cell_b.number_format = NUM_FMT_CF
    else:
        cell_b.value = 'ดูรายละเอียด'
        cell_b.font = font(color='7F7F7F', italic=True, size=9)

    cell_a.font = font(size=9)
    cell_b.font = font(size=9, color='000000')
    cell_c.font = font(size=9, color='5A5A5A', italic=True)
    for cell in [cell_a, cell_b, cell_c]:
        cell.border = border_thin()
        cell.alignment = align('left', 'center')
    cell_b.alignment = align('right', 'center')

    if 'Operating' in activity:
        op_rows.append(r)
    elif 'Financing' in activity:
        fin_rows.append(r)
    elif 'Investing' in activity:
        inv_rows.append(r)

    ws2.row_dimensions[r].height = 15
    r += 1

r += 1

# Operating Total
def add_subtotal(label, rows, row_idx):
    ws2.cell(row=row_idx, column=1, value=label).font = font(bold=True, size=10)
    if rows:
        formula = '+'.join([f'B{x}' for x in rows if section_summary_rows[[s[1] for s in section_summary_rows].index(x)][2] != 0] if False else [f'B{x}' for x in rows])
        cell = ws2.cell(row=row_idx, column=2, value=f'=SUM({",".join([f"B{x}" for x in rows])})')
        cell.number_format = NUM_FMT_CF
        cell.font = font(bold=True, size=10)
    for col in [1, 2, 3]:
        ws2.cell(row=row_idx, column=col).fill = fill(COL_TOTAL_BG)
        ws2.cell(row=row_idx, column=col).border = border_thin()
        ws2.cell(row=row_idx, column=col).alignment = align('right' if col == 2 else 'left', 'center')
    ws2.row_dimensions[row_idx].height = 17

if op_rows:
    add_subtotal('รวมผลกระทบ — Operating Activities', op_rows, r); r += 1
if inv_rows:
    add_subtotal('รวมผลกระทบ — Investing Activities', inv_rows, r); r += 1
if fin_rows:
    add_subtotal('รวมผลกระทบ — Financing Activities', fin_rows, r); r += 1

r += 1

# Legend
ws2.merge_cells(f'A{r}:C{r}')
ws2[f'A{r}'] = 'หมายเหตุ: ตัวเลขสีน้ำเงินใน Working Paper = กรอกเอง (Opening Balance)  |  สีเขียว = จาก TB  |  เซลล์สีเหลือง = Intercompany (IC) — ถูก Eliminate ใน CONSO'
ws2[f'A{r}'].font = font(size=8, color='444444', italic=True)
ws2[f'A{r}'].fill = fill('F9F9F9')
ws2[f'A{r}'].alignment = align('left', 'center', wrap=True)
ws2.row_dimensions[r].height = 28

# ── Sheet 3: Instructions ─────────────────────────────────────────────────────
ws3 = wb.create_sheet('Instructions')
ws3.column_dimensions['A'].width = 80
instructions = [
    ('วิธีใช้ Working Paper Template นี้', True, 12, COL_HDR_BG, COL_HDR_FG),
    ('', False, 10, 'FFFFFF', '000000'),
    ('STEP 1 — กรอก Opening Balance (คอลัมน์ E สีน้ำเงิน)', True, 10, COL_SEC_BG, COL_SEC_FG),
    ('  กรอกยอดยกมาต้นงวด (เช่น 31 Dec 2023) สำหรับทุก Account', False, 9, 'FFFFFF', '000000'),
    ('  ถ้ามี TB งวดก่อนหน้า สามารถ paste ค่าตรงๆ ได้เลย', False, 9, 'FFFFFF', '000000'),
    ('', False, 9, 'FFFFFF', '000000'),
    ('STEP 2 — ตรวจสอบ Closing Balance (คอลัมน์ F สีเขียว)', True, 10, COL_SEC_BG, COL_SEC_FG),
    ('  คอลัมน์ F ดึงมาจาก TB ที่ให้มา (CONSO) โดยอัตโนมัติ', False, 9, 'FFFFFF', '000000'),
    ('  ถ้าตัวเลขไม่ตรงกับ TB ให้แจ้งทีม Accounting', False, 9, 'FFFFFF', '000000'),
    ('', False, 9, 'FFFFFF', '000000'),
    ('STEP 3 — ระบบคำนวณ Change และ CF Impact ให้อัตโนมัติ', True, 10, COL_SEC_BG, COL_SEC_FG),
    ('  Change = Closing - Opening (คอลัมน์ G)', False, 9, 'FFFFFF', '000000'),
    ('  CF Impact = Change * Sign (คอลัมน์ H) — ดูสูตรแต่ละ Section', False, 9, 'FFFFFF', '000000'),
    ('', False, 9, 'FFFFFF', '000000'),
    ('STEP 4 — ดูสรุปใน Sheet "CF Summary (WC)"', True, 10, COL_SEC_BG, COL_SEC_FG),
    ('  Sheet นี้รวม CF Impact แยกตาม Operating / Investing / Financing', False, 9, 'FFFFFF', '000000'),
    ('', False, 9, 'FFFFFF', '000000'),
    ('สัญลักษณ์สีที่ใช้', True, 10, '1F3864', COL_HDR_FG),
    ('  ตัวเลขสีน้ำเงิน = กรอกเอง (Hardcoded Input)', False, 9, 'FFFFFF', '0000FF'),
    ('  ตัวเลขสีเขียว = ดึงจาก TB', False, 9, 'FFFFFF', '006100'),
    ('  ตัวเลขสีดำ = สูตรคำนวณ', False, 9, 'FFFFFF', '000000'),
    ('  เซลล์สีเหลือง = Intercompany (IC) — ถูก Eliminate ใน CONSO', False, 9, COL_WARN_BG, '7B3F00'),
    ('  เซลล์สีเขียวอ่อน = แถว Total', False, 9, COL_TOTAL_BG, '000000'),
    ('', False, 9, 'FFFFFF', '000000'),
    ('TFRS 16 — ข้อควรระวัง', True, 10, '1F3864', COL_HDR_FG),
    ('  Section 4 (ROU Asset) และ Section 8 (Lease Liability) = TFRS 16', False, 9, 'FFFFFF', '000000'),
    ('  CF Impact ของ 2 Section นี้ถูกระบุว่า N/A เพราะต้องแยก:', False, 9, 'FFFFFF', '000000'),
    ('    - ค่าเสื่อม ROU Asset = Add-back ใน Operating (Non-cash)', False, 9, 'FFFFFF', '000000'),
    ('    - ชำระเงินต้น Lease Liability = Financing Outflow', False, 9, 'FFFFFF', '000000'),
    ('    - ดอกเบี้ย Lease = Operating หรือ Financing ตาม Policy บริษัท', False, 9, 'FFFFFF', '000000'),
    ('  ให้ขอ TFRS 16 Schedule จากฝ่ายบัญชีเพิ่มเติม', False, 9, 'FFFFFF', 'CC0000'),
]
for i, (text, bold, size, bg, fg) in enumerate(instructions, start=1):
    cell = ws3.cell(row=i, column=1, value=text)
    cell.font = Font(name=FONT_NAME, bold=bold, size=size, color=fg)
    cell.fill = fill(bg)
    cell.alignment = align('left', 'center', wrap=True)
    ws3.row_dimensions[i].height = 16 if text else 8

wb.save('Working_Paper_Template.xlsx')
print('Done: Working_Paper_Template.xlsx')
print(f'Total sections: {len(SECTIONS)}')
total_acc = sum(len(v) for v in section_data.values())
print(f'Total accounts mapped: {total_acc}')
print(f'Unmatched BS rows: {len(unmatched_bs)}')
