import openpyxl
import csv
from collections import Counter

print("Reading TB Mar 2026...")
wb = openpyxl.load_workbook("sample-data/TB by Entity&Conso_31.03.26.xlsx", read_only=True, data_only=True)
ws = wb.active

accounts_mar26 = {}
for row in ws.iter_rows(values_only=True):
    if row[1] == "ACCOUNT":
        continue
    if not row[2]:
        continue
    acc_code = str(int(row[2])) if isinstance(row[2], float) else str(row[2])
    if acc_code not in accounts_mar26:
        accounts_mar26[acc_code] = {
            "bs_class": row[0] or "",
            "account_name": row[3] or "",
            "sub_caption": row[4] or "",
            "fs_caption": row[5] or "",
            "acg_mar26": row[6] or 0,
            "hmw_mar26": row[10] or 0,
            "clik_mar26": row[14] or 0,
            "conso_mar26": row[19] or 0,
        }
wb.close()
print(f"  Found {len(accounts_mar26)} accounts in Mar 2026")

print("Reading TB Dec 2025...")
wb2 = openpyxl.load_workbook("sample-data/TB_YE2568_31Dec2025.xlsx", read_only=True, data_only=True)
ws2 = wb2.active

dec25_data = {}
for row in ws2.iter_rows(values_only=True):
    if row[1] == "ACCOUNT":
        continue
    if not row[2]:
        continue
    acc_code = str(int(row[2])) if isinstance(row[2], float) else str(row[2])
    if acc_code not in dec25_data:
        dec25_data[acc_code] = {
            "acg_dec25": row[6] or 0,
            "hmw_dec25": row[10] or 0,
            "clik_dec25": row[14] or 0,
            "conso_dec25": row[19] or 0,
        }
wb2.close()
print(f"  Found {len(dec25_data)} accounts in Dec 2025")

all_codes = sorted(set(list(accounts_mar26.keys()) + list(dec25_data.keys())))
print(f"\nTotal unique accounts across both periods: {len(all_codes)}")

cls_count = Counter(accounts_mar26.get(c, {}).get("bs_class", "Unknown") for c in all_codes)
print("By class:", dict(cls_count))

# Classify by account code prefix for CF mapping
def get_cf_category(code, bs_class, fs_cap):
    prefix = code[:3]
    prefix4 = code[:4]
    if prefix in ("111",):
        return "Cash & Bank"
    elif prefix in ("112", "113", "114"):
        return "Trade Receivables / AR"
    elif prefix in ("115", "116", "117"):
        return "Other Current Assets"
    elif prefix in ("121", "122", "123"):
        return "Inventories"
    elif prefix in ("130", "131", "132"):
        return "Investment"
    elif prefix in ("140", "141", "142", "143"):
        return "Fixed Assets / PPE"
    elif prefix in ("150", "151", "152"):
        return "Intangible Assets"
    elif prefix in ("210", "211", "212"):
        return "Trade Payables / AP"
    elif prefix in ("213", "214", "215", "216", "217", "218", "219", "220", "221", "222", "223"):
        return "Other Current Liabilities"
    elif prefix in ("230", "231", "232"):
        return "Long-term Debt"
    elif prefix in ("310", "311", "312", "313", "314", "315"):
        return "Equity / Share Capital"
    elif prefix in ("316",):
        return "Retained Earnings"
    elif prefix in ("410", "411", "412", "413", "414", "415", "416", "417", "418", "419"):
        return "Revenue / Income"
    elif prefix in ("510", "511", "512", "513", "514", "515", "516", "517", "518", "519", "520", "521"):
        return "Cost of Sales / COGS"
    elif prefix in ("540", "541", "542", "543", "544", "545", "546"):
        return "Selling & Distribution Expense"
    elif prefix in ("560", "561", "562", "563", "564", "565", "566"):
        return "G&A Expense"
    elif prefix in ("580", "581", "582"):
        return "Finance Cost"
    elif prefix in ("590", "591", "592"):
        return "Other Income / Expense"
    else:
        return f"Other ({bs_class})"

# Write COA CSV
output_file = "data/COA_from_TB.csv"
fieldnames = [
    "account_code", "account_name_th", "sub_caption_th", "fs_caption_th",
    "bs_class", "cf_category",
    "acg_dec25", "hmw_dec25", "clik_dec25", "conso_dec25",
    "acg_mar26", "hmw_mar26", "clik_mar26", "conso_mar26",
    "companies_active"
]

rows = []
for code in all_codes:
    a = accounts_mar26.get(code, {})
    d = dec25_data.get(code, {})

    bs_class = a.get("bs_class", "")
    fs_cap = a.get("fs_caption", "")
    cf_cat = get_cf_category(code, bs_class, fs_cap)

    acg_m = a.get("acg_mar26", 0)
    hmw_m = a.get("hmw_mar26", 0)
    clik_m = a.get("clik_mar26", 0)

    companies = []
    if acg_m != 0 or d.get("acg_dec25", 0) != 0:
        companies.append("ACG")
    if hmw_m != 0 or d.get("hmw_dec25", 0) != 0:
        companies.append("HMW")
    if clik_m != 0 or d.get("clik_dec25", 0) != 0:
        companies.append("CLIK")

    rows.append({
        "account_code": code,
        "account_name_th": a.get("account_name", ""),
        "sub_caption_th": a.get("sub_caption", ""),
        "fs_caption_th": a.get("fs_caption", ""),
        "bs_class": bs_class,
        "cf_category": cf_cat,
        "acg_dec25": d.get("acg_dec25", 0),
        "hmw_dec25": d.get("hmw_dec25", 0),
        "clik_dec25": d.get("clik_dec25", 0),
        "conso_dec25": d.get("conso_dec25", 0),
        "acg_mar26": acg_m,
        "hmw_mar26": hmw_m,
        "clik_mar26": clik_m,
        "conso_mar26": a.get("conso_mar26", 0),
        "companies_active": ",".join(companies) if companies else "-",
    })

with open(output_file, "w", newline="", encoding="utf-8-sig") as f:
    writer = csv.DictWriter(f, fieldnames=fieldnames)
    writer.writeheader()
    writer.writerows(rows)

print(f"\nCOA written to {output_file} ({len(rows)} accounts)")

# Summary by CF category
cat_count = Counter(r["cf_category"] for r in rows)
print("\nCF Category summary:")
for cat, cnt in sorted(cat_count.items(), key=lambda x: -x[1]):
    print(f"  {cat}: {cnt} accounts")
