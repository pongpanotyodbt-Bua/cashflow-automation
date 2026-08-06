"""
Build TB Sample for 31/12/2568 (Year End 2025)
Based on Consolidated Financial Statements
Same format as TB by Entity&Conso.xlsx
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def fix(s):
    if isinstance(s, str):
        try: return s.encode('latin1').decode('cp874')
        except: return s
    return s
def num(x): return x if isinstance(x,(int,float)) else 0

# ── Load current TB for account master (acc_no, acc_name, sub, fs, cat) ──────
tb_cur = openpyxl.load_workbook('sample-data/TB by Entity&Conso.xlsx', read_only=True, data_only=True)
ws_cur = tb_cur['Sheet1']

acc_master = []
for i, row in enumerate(ws_cur.iter_rows(values_only=True)):
    if i == 0: continue
    if row[1] is None and row[3] is None: continue
    acc_no = row[2]
    if not acc_no: continue
    acc_master.append({
        'cat':        row[0],
        'account':    row[1],
        'acc_no':     acc_no,
        'acc_name':   fix(row[3]) if row[3] else '',
        'sub':        fix(row[4]) if row[4] else '',
        'fs':         fix(row[5]) if isinstance(row[5], str) else (str(row[5]) if row[5] else ''),
        'acg_cur':    num(row[6]),
        'hmw_cur':    num(row[10]),
        'clik_cur':   num(row[14]),
        'conso_cur':  num(row[19]),
        'conso_adj':  num(row[18]),
    })

# ── FS Caption → 31/12/2568 CONSO balance (from Financial Statements 2568) ──
# Asset = positive, Liability/Equity = negative (normal balance)
FS_BALANCE_2568 = {
    'เงินสดและรายการเทียบเท่าเงินสด':          173743169,
    'ลูกหนี้การค้า':                             21606892,
    'ลูกหนี้หมุนเวียนอื่น':                     13437598,
    'สินค้าคงเหลือ':                             74124409,
    'สินทรัพย์หมุนเวียนอื่น':                    1496906,
    'สินทรัพย์ภาษีเงินได้ของงวดปัจจุบัน':        0,
    'เงินลงทุนในบริษัทย่อย':                     0,
    'ลูกหนี้ไม่หมุนเวียนอื่น':                   4489980,
    'อสังหาริมทรัพย์เพื่อการลงทุน':             144164368,
    'ที่ดิน อาคารและอุปกรณ์ ':                  558287983,
    'สินทรัพย์สิทธิการใช้':                     141912925,
    'สินทรัพย์ไม่มีตัวตน':                       6457013,
    'สินทรัพย์ภาษีเงินได้รอการตัดบัญชี':        14845239,
    'สินทรัพย์ไม่หมุนเวียนอื่น':                 9928613,
    'เงินให้กู้ยืมระยะยาวแก่บริษัทย่อย':         0,
    'เงินให้กู้ยืมระยะสั้นจากสถาบันการเงิน':   -120000000,
    'เงินเบิกเกินบัญชีและเงินกู้ยืมระยะสั้นจากสถาบันการเงิน': -120000000,
    'เจ้าหนี้การค้า':                           -87097100,
    'เจ้าหนี้หมุนเวียนอื่น':                    -24138701,
    'ที่ถึงกำหนดชำระภายในหนึ่งปี':              -10510744,
    'ภาษีเงินได้นิติบุคคลค้างจ่าย':              -5996908,
    'หนี้สินหมุนเวียนอื่น':                      -5497209,
    'เงินกู้ยืมระยะยาวจากสถาบันการเงิน':          0,
    'หนี้สินตามสัญญาเช่า':                     -157619708,
    'ผลประโยชน์พนักงาน':                        -10858270,
    'ประมาณการหนี้สินไม่หมุนเวียนอื่น':          -6213079,
    'ทุนจดทะเบียน':                            -312000000,
    'ทุนที่ออกและชำระแล้ว':                    -300000000,
    'ส่วนเกินมูลค่าหุ้นสามัญ':                 -137109509,
    'ส่วนเกินทุนจากการจัดโครงสร้างธุรกิจ':     -130891299,
    'ใบสำคัญแสดงสิทธิที่จะซื้อหุ้น':            -16113897,
    'ทุนสำรองตามกฎหมาย':                        -8378303,
    'กำไรสะสมยังไม่ได้จัดสรร':                -141966488,
    'ส่วนได้เสียที่ไม่มีอำนาจควบคุม':            -2103880,
    # PL accounts (net for full year 2568)
    'รายได้จากการขาย':                        -1320815808,
    'รายได้จากบริการ':                          -24126468,
    'รายได้อื่น':                               -13250158,
    'ต้นทุนขาย':                              1079265810,
    'ต้นทุนการให้บริการ':                       56360501,
    'ค่าใช้จ่ายในการบริหาร':                   153605579,
    'ต้นทุนทางการเงิน':                         14962373,
    'ค่าใช้จ่ายภาษีเงินได้':                    16215478,
}

# ── Assign CONSO balance per account based on FS Caption matching ─────────────
def get_conso_2568(acc):
    fs = acc['fs'].strip()
    cat = str(acc['cat'])
    cur_conso = acc['conso_cur']

    # Direct FS match
    for key, val in FS_BALANCE_2568.items():
        if key in fs or fs in key:
            return val, True

    # Fallback: scale from current TB proportionally if same FS group
    # For accounts with no direct match, use current TB ratio
    return cur_conso, False

# ── Build entity splits (ACG/HMW/CLIK) based on current TB proportions ───────
def split_entities(acc, conso_val):
    cur_conso = acc['conso_cur']
    if cur_conso == 0:
        # Distribute proportionally by current entity values
        acg = acc['acg_cur']
        hmw = acc['hmw_cur']
        clik = acc['clik_cur']
        total = acg + hmw + clik
        if total == 0:
            return conso_val, 0, 0  # All to ACG if zero
        acg_new  = round(conso_val * acg / total, 2) if total else conso_val
        hmw_new  = round(conso_val * hmw / total, 2) if total else 0
        clik_new = round(conso_val - acg_new - hmw_new, 2)
        return acg_new, hmw_new, clik_new
    else:
        ratio = conso_val / cur_conso
        acg_new  = round(acc['acg_cur']  * ratio, 2)
        hmw_new  = round(acc['hmw_cur']  * ratio, 2)
        clik_new = round(conso_val - acg_new - hmw_new, 2)
        return acg_new, hmw_new, clik_new

# ── Group accounts by FS Caption to distribute total ──────────────────────────
from collections import defaultdict
fs_groups = defaultdict(list)
for acc in acc_master:
    fs_groups[acc['fs'].strip()].append(acc)

# ── Calculate per-account CONSO balance ──────────────────────────────────────
results = []
fs_used_totals = defaultdict(float)

for acc in acc_master:
    fs = acc['fs'].strip()
    cur_conso = acc['conso_cur']
    cur_conso_sum = sum(a['conso_cur'] for a in fs_groups[fs])

    # Find FS balance
    fs_total = None
    for key, val in FS_BALANCE_2568.items():
        if key in fs or fs in key:
            fs_total = val
            break

    if fs_total is not None and cur_conso_sum != 0:
        # Proportional split within same FS group
        weight = cur_conso / cur_conso_sum if cur_conso_sum != 0 else 0
        conso_new = round(fs_total * weight, 2)
    elif fs_total is not None:
        conso_new = fs_total
    else:
        # No match — keep same as current (will be marked as unverified)
        conso_new = cur_conso

    # Entity split
    acg_new, hmw_new, clik_new = split_entities(acc, conso_new)
    conso_adj_new = acc['conso_adj']

    results.append({
        **acc,
        'acg_new':       acg_new,
        'hmw_new':       hmw_new,
        'clik_new':      clik_new,
        'conso_adj_new': conso_adj_new,
        'conso_new':     conso_new,
        'fs_matched':    fs_total is not None,
    })

# ── Write Excel ───────────────────────────────────────────────────────────────
wb_out = openpyxl.Workbook()
ws_out = wb_out.active
ws_out.title = 'Sheet1'

# Header row (same as original)
headers = [None, 'ACCOUNT', 'ACCOUNT NO.', 'ACCOUNT NAME', 'SUB CAPTION', 'FS CAPTION',
           'ACG', 'ACG DR', 'ACG CR', 'ACG conso',
           'HMW', 'HMW DR', 'HMW CR', 'HMW conso',
           'CLIK', 'CLIK DR', 'CLIK CR', 'CLIK conso',
           'Conso adj.', 'CONSO']

# Style
hdr_font  = Font(name='Arial', bold=True, color='FFFFFF', size=10)
hdr_fill  = PatternFill('solid', fgColor='1F3864')
hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
thin      = Side(style='thin', color='CCCCCC')
bdr       = Border(left=thin, right=thin, top=thin, bottom=thin)

ws_out.row_dimensions[1].height = 28
for col_idx, h in enumerate(headers, start=1):
    cell = ws_out.cell(row=1, column=col_idx, value=h)
    cell.font = hdr_font
    cell.fill = hdr_fill
    cell.alignment = hdr_align
    cell.border = bdr

# Col widths
widths = [4, 12, 14, 45, 35, 45, 16, 14, 14, 16, 16, 14, 14, 16, 16, 14, 14, 16, 16, 16]
for i, w in enumerate(widths, start=1):
    ws_out.column_dimensions[get_column_letter(i)].width = w

# Data rows
cat_fills = {
    'Asset':       PatternFill('solid', fgColor='EBF5FB'),
    'Liabilities': PatternFill('solid', fgColor='FDEDEC'),
    'Equity':      PatternFill('solid', fgColor='EAFAF1'),
    'PL':          PatternFill('solid', fgColor='FEF9E7'),
}
unmatched_fill = PatternFill('solid', fgColor='FFF3CD')
num_fmt  = '#,##0.00_);(#,##0.00);"-"'
data_align_l = Alignment(horizontal='left',  vertical='center')
data_align_r = Alignment(horizontal='right', vertical='center')
data_font    = Font(name='Arial', size=9)
warn_font    = Font(name='Arial', size=9, color='E67E22')

for r_idx, acc in enumerate(results, start=2):
    cat = str(acc['cat']) if acc['cat'] else ''
    bg  = cat_fills.get(cat, PatternFill('solid', fgColor='FFFFFF'))
    if not acc['fs_matched']:
        bg = unmatched_fill

    row_data = [
        cat,
        acc['account'],
        acc['acc_no'],
        acc['acc_name'],
        acc['sub'],
        acc['fs'],
        acc['acg_new'],
        0,  # DR
        0,  # CR
        acc['acg_new'],
        acc['hmw_new'],
        0,
        0,
        acc['hmw_new'],
        acc['clik_new'],
        0,
        0,
        acc['clik_new'],
        acc['conso_adj_new'],
        acc['conso_new'],
    ]

    for col_idx, val in enumerate(row_data, start=1):
        cell = ws_out.cell(row=r_idx, column=col_idx, value=val)
        cell.fill = bg
        cell.border = bdr
        if col_idx <= 6:
            cell.alignment = data_align_l
            cell.font = warn_font if not acc['fs_matched'] else data_font
        else:
            cell.alignment = data_align_r
            cell.font = data_font
            if isinstance(val, (int, float)):
                cell.number_format = num_fmt

    ws_out.row_dimensions[r_idx].height = 14

# ── Summary sheet ─────────────────────────────────────────────────────────────
ws_sum = wb_out.create_sheet('Summary & Verification')
ws_sum.column_dimensions['A'].width = 50
ws_sum.column_dimensions['B'].width = 22
ws_sum.column_dimensions['C'].width = 22

sum_headers = ['รายการ', 'FS งบการเงิน 31/12/2568', 'TB Sample 31/12/2568']
for col_idx, h in enumerate(sum_headers, start=1):
    c = ws_sum.cell(row=1, column=col_idx, value=h)
    c.font = hdr_font; c.fill = hdr_fill; c.alignment = hdr_align

check_items = [
    ('เงินสดและรายการเทียบเท่าเงินสด', 173743169),
    ('ลูกหนี้การค้า', 21606892),
    ('ลูกหนี้หมุนเวียนอื่น', 13437598),
    ('สินค้าคงเหลือ', 74124409),
    ('สินทรัพย์หมุนเวียนอื่น', 1496906),
    ('อสังหาริมทรัพย์เพื่อการลงทุน', 144164368),
    ('ที่ดิน อาคารและอุปกรณ์', 558287983),
    ('สินทรัพย์สิทธิการใช้ (ROU)', 141912925),
    ('สินทรัพย์ภาษีเงินได้รอตัดบัญชี', 14845239),
    ('รวมสินทรัพย์', 1164495095),
    ('เจ้าหนี้การค้า', -87097100),
    ('เงินกู้ยืมระยะสั้น', -120000000),
    ('หนี้สินตามสัญญาเช่า (รวม)', -168130452),
    ('ภาษีเงินได้นิติบุคคลค้างจ่าย', -5996908),
    ('ผลประโยชน์พนักงาน', -10858270),
    ('รวมหนี้สิน', -427931719),
]

fs_caption_totals = defaultdict(float)
for acc in results:
    fs_caption_totals[acc['fs'].strip()] += acc['conso_new']

for r_idx, (label, fs_val) in enumerate(check_items, start=2):
    tb_val = 0
    for key, val in fs_caption_totals.items():
        if label[:8] in key or key[:8] in label:
            tb_val += val

    c1 = ws_sum.cell(row=r_idx, column=1, value=label)
    c2 = ws_sum.cell(row=r_idx, column=2, value=fs_val)
    c3 = ws_sum.cell(row=r_idx, column=3, value=round(tb_val, 0))
    diff = abs(fs_val - round(tb_val, 0))
    ok = diff < abs(fs_val) * 0.01 if fs_val != 0 else diff < 1000

    for c in [c1, c2, c3]:
        c.font = Font(name='Arial', size=9, color='006100' if ok else 'CC0000')
        c.fill = PatternFill('solid', fgColor='EAF7EA' if ok else 'FDEDEC')
        c.border = bdr
        c.number_format = num_fmt
        c.alignment = data_align_r
    c1.alignment = data_align_l

# Note
ws_sum.cell(row=len(check_items)+3, column=1,
    value='หมายเหตุ: เซลล์สีเหลืองใน Sheet1 = FS Caption ไม่ match โดยตรง — ใช้ค่าเดิมจาก Q1/2569 TB').font = Font(name='Arial', size=8, color='7B3F00', italic=True)

wb_out.save('sample-data/TB_YE2568_31Dec2025.xlsx')
print('Done: sample-data/TB_YE2568_31Dec2025.xlsx')
matched = sum(1 for r in results if r['fs_matched'])
print(f'Accounts matched to FS: {matched}/{len(results)}')
print(f'Unmatched (yellow): {len(results)-matched}')
