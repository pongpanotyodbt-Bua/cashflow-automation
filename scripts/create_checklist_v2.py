import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

csv_path = 'C:\\Users\\pongpanotyod.b\\Desktop\\Cashflow Management\\Check_List_Request_Data_v2.csv'
output_path = 'C:\\Users\\pongpanotyod.b\\Desktop\\Cashflow Management\\Check_List_Request_Data_v2.xlsx'

df = pd.read_csv(csv_path)

wb = Workbook()
wb.remove(wb.active)

header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11, name="Arial")
subheader_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
subheader_font = Font(bold=True, color="000000", size=10, name="Arial")
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='top', wrap_text=True)

dept_colors = {
    'Finance': 'FFE699',
    'AR': 'C5E0B4',
    'Accounting': 'B4C7E7',
    'Treasury': 'F8CBAD'
}

def create_data_sheet(sheet_name, dept_name):
    ws = wb.create_sheet(sheet_name)

    headers = ['#', 'ชื่อข้อมูล', 'ชื่ออื่นๆ', 'วัตถุประสงค์', 'รูปแบบ', 'ความถี่', 'หมายเหตุ', 'ได้รับ']

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border

    data_rows = df[df['แผนก'] == dept_name].copy()
    excel_row = 2

    for _, row in data_rows.iterrows():
        item_num = row['#']

        if pd.isna(item_num) or (isinstance(item_num, float) and pd.isna(item_num)):
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=excel_row, column=col)
                cell.value = row[headers[col-1]] if col == 2 else ""
                cell.font = subheader_font
                cell.fill = subheader_fill
                cell.border = border
                cell.alignment = left_align
        else:
            try:
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=excel_row, column=col)
                    cell.value = row[header]
                    cell.border = border
                    cell.alignment = center_align if col == 1 or col == 8 else left_align
                    if col == 8:
                        cell.font = Font(size=12)
                excel_row += 1
            except (ValueError, TypeError):
                pass

    ws.column_dimensions['A'].width = 4
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 28
    ws.column_dimensions['D'].width = 33
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 33
    ws.column_dimensions['H'].width = 8

    ws.freeze_panes = 'A2'
    return ws

create_data_sheet('Finance', 'Finance')
create_data_sheet('AR', 'AR')
create_data_sheet('Accounting', 'Accounting')
create_data_sheet('Treasury', 'Treasury')

summary = wb.create_sheet('Summary', 0)

summary['A1'] = 'CHECK LIST ขอข้อมูล - ระบบ Cashflow Management'
summary['A1'].font = Font(bold=True, size=14, color="FFFFFF", name="Arial")
summary['A1'].fill = PatternFill(start_color="203864", end_color="203864", fill_type="solid")
summary['A1'].alignment = center_align
summary.merge_cells('A1:D1')

summary['A2'] = 'Version 2 (ตัด Forecast, เพิ่ม Assets/Cost/Loan) | 38 รายการ | 4 แผนก'
summary['A2'].font = Font(size=10, color="666666", name="Arial")
summary.merge_cells('A2:D2')

summary['A4'] = 'แผนก'
summary['B4'] = 'จำนวนรายการ'
summary['C4'] = 'ลำดับที่'
summary['D4'] = 'สถานะ'

for col in ['A', 'B', 'C', 'D']:
    summary[f'{col}4'].font = header_font
    summary[f'{col}4'].fill = header_fill
    summary[f'{col}4'].border = border
    summary[f'{col}4'].alignment = center_align

summary_data = [
    ('💰 Finance', 14, '1-14', '☐'),
    ('💳 AR (Account Receivable)', 10, '15-24', '☐'),
    ('📈 Accounting', 8, '25-32', '☐'),
    ('🏦 Treasury', 6, '33-38', '☐')
]

for idx, (dept, count, rng, status) in enumerate(summary_data, 5):
    summary[f'A{idx}'] = dept
    summary[f'B{idx}'] = count
    summary[f'C{idx}'] = rng
    summary[f'D{idx}'] = status
    for col in ['A', 'B', 'C', 'D']:
        summary[f'{col}{idx}'].border = border
        summary[f'{col}{idx}'].alignment = center_align

summary['A9'] = 'รวมทั้งหมด'
summary['B9'] = 38
summary['A9'].font = Font(bold=True, name="Arial")
summary['B9'].font = Font(bold=True, name="Arial")
for col in ['A', 'B']:
    summary[f'{col}9'].border = border
    summary[f'{col}9'].alignment = center_align
summary['A9'].fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
summary['B9'].fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")

summary['A11'] = 'การเปลี่ยนแปลง:'
summary['A11'].font = Font(bold=True, size=11, name="Arial")
summary['A12'] = '✓ ตัด Forecast Cash ออก (-4 รายการ)'
summary['A13'] = '✓ เพิ่ม Fixed Assets, Cost Structure, Loan & Debt (+9 รายการ)'
summary['A14'] = '✓ เพิ่ม Treasury Department (จัดการ Loan & Asset)'
summary['A15'] = '✓ ยังคงต้อง Indirect Method สำหรับคำนวณ Cash Flow'

summary.column_dimensions['A'].width = 50
summary.column_dimensions['B'].width = 15
summary.column_dimensions['C'].width = 12
summary.column_dimensions['D'].width = 10

wb.save(output_path)
print(f"✓ สร้าง Excel Check list v2 เรียบร้อย: Check_List_Request_Data_v2.xlsx")
print(f"  - Finance: 14 รายการ")
print(f"  - AR: 10 รายการ")
print(f"  - Accounting: 8 รายการ")
print(f"  - Treasury: 6 รายการ")
print(f"  - รวม: 38 รายการ")
